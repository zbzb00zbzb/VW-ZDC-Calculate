from openpyxl import Workbook
from openpyxl import load_workbook
import operator
from array import array
import re
#from Excel import *
from bs4 import BeautifulSoup
import bs4
import os,os.path
import shutil
import sys
import PyQt5.QtCore
from PyQt5.QtCore import pyqtSignal, QObject
from PyQt5.QtWidgets import (QMainWindow, QTextEdit,
                             QAction, QFileDialog, QApplication, QLabel, QLineEdit, QGridLayout, QWidget, QPushButton,QMessageBox)
from PyQt5.QtGui import QIcon

class Kopp_Analyse_and_ZDC_Calculate_class(QObject):
    trigger_output_k = PyQt5.QtCore.pyqtSignal(str)

    def __init__(self):
        print('Kopp_Analyse_and_ZDC_Calculate_class chushihua ')

        super().__init__()

        self.initUI()

    def initUI(self):
        self.kopp_file_path=' '
        self.ZDC_file_list=[]

    def example(self,text):
        print('example')
        self.trigger_output.emit(text)

    def Kopp_Analyse_and_ZDC_Calculate(self,kopp_file_path,option_flag):
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/20180115 Polo NF BF KOPP SVMP update KW01.xlsx")
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/20180313 T-Cross KOPP_After LF.xlsx")
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/2017-11-28_KOPP_Tharu LF - 副本.xlsx")
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/Model Q KOPP 20170818 after SZE .xlsx")
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/20171123 Lavida NF_KOPP_SE_UG0 & 6LZ update.xlsx")
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/20180209 Karoq KOPP_V7 1 for CPC.xlsx")
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/KODIAQ KOPP （MY19 2018 KW39）V3 0 (1 4T2 0T High C6) 20171225.xlsx")
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/Superb MP18&Business KOPP 20170322 - 2.0 working.xlsx")
        #wb = load_workbook("C:/Users/SVW/Desktop/Kopp-all/20180309 NMS PHEV Kopp incl Fleet based on LF Update_Signature.xlsx")
        #kopp_file_path="C:/Users/SVW/Desktop/Kopp-all/20180308 NMS NF Kopp -LF PR Update with Update Signature.xlsx"
        # kopp_file_path="C:/Users/SVW/Desktop/Kopp-all/20180310 Octavia Combi 1.5L & 1.2T C6 KOPP_V2.4.xlsx"

        Error_info='Error_0'
        if len(kopp_file_path)<4:
            Error_info ='kopp文件路径错误，请重新加载kopp文件'
            return (Error_info, [], [], [])
        wb = load_workbook(kopp_file_path)
        print('开始计算kopp文件')
        sheets_name = wb.get_sheet_names()
        #print(sheets_name)
        row_length_max_sheet_index=0
        row_length_max=0
        Equipmentline_list=[]
        PR_FAM_list_list=[]
        PR_Num_list_list=[]
        #PR_FAM_list_list = [[] for i in range(20)]
        #PR_Num_list_list=[[] for i in range(20)]

        for i in range(len(sheets_name)):
            ws = wb.get_sheet_by_name(sheets_name[i])#只导入Dipon页
            ws_rows_len = ws.max_row
            ws_columns_len = ws.max_column
            if (ws_rows_len>row_length_max) and (250<ws_rows_len<1000):
                row_length_max=ws_rows_len
                row_length_max_sheet_index=i

        ####################################################################################################################
        ws = wb.get_sheet_by_name(sheets_name[row_length_max_sheet_index])#只导入Dipon页
        ws_rows_len = ws.max_row         #文件行数
        ws_columns_len = ws.max_column    #文件列数
        temp="Kopp表所在的sheet的名称: "+str(sheets_name[row_length_max_sheet_index])+"， 共有"+str(row_length_max)+'行.'
        self.trigger_output_k.emit(temp)
        Option_flag=option_flag# 1表示要带选装包，0表示不带选装包

        start_row = 1#起始行数
        start_column = 1#起始烈数
        PR_data = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]#初始化PR数据
        equip_lines_data = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]#初始化装备线数据
        PR_FAM_column=0#FAM所在列
        PR_Num_column=0#PRN所在列

        for column in range(start_column, ws_columns_len+1):#遍历所有列
          for row in range(start_row, ws_rows_len + 1):#遍历所有行
            if len(str(ws.cell(row,column).value))==3:
                PR_data[column-1]=PR_data[column-1]+1#计数
            elif len(str(ws.cell(row,column).value))==1:
                if  str(ws.cell(row,column).value)=='X' or str(ws.cell(row,column).value)=='O'or str(ws.cell(row,column).value)=='o'or str(ws.cell(row,column).value)=='x'or str(ws.cell(row,column).value)=='0'or str(ws.cell(row,column).value)=='●':
                    equip_lines_data[column-1]=equip_lines_data[column-1]+1#计数
                else:
                    pass
            else :
                pass

        PR_Num_column=PR_data.index(max(PR_data))+1
        Max_1=max(PR_data)
        PR_data[PR_Num_column-1]=0
        Max_2=max(PR_data)
        PR_FAM_column=PR_data.index(max(PR_data))+1

        print(equip_lines_data)
        equip_lines=[]#初始化装备线所在的列
        for i,j in enumerate( equip_lines_data):#长度最大的几列是装备线
            if((int(max(equip_lines_data))-30)<int(j)<(int(max(equip_lines_data)+15))):
                    equip_lines.append(i+1)#将列数存入表中
        if(Max_2<(Max_1/4)):
            PR_FAM_column=PR_Num_column#将FAM和PRN定义为同一列

        lines_row_length=[]
        for j in list(range(10)):
            length=0
            for i in equip_lines:#遍历每行
                length=length+len(str(ws.cell(j+1, i).value))#计数
            lines_row_length.append(length)#保存每行长度
        equip_lines_row = lines_row_length.index(max(lines_row_length))+1
        print('equip_lines_row:',equip_lines_row)
        '''
        self.trigger_output_k.emit( "PR_FAM所在的列数：")
        self.trigger_output_k.emit(str(PR_FAM_column))#FAM所在列
        self.trigger_output_k.emit( "PR_Num所在的列数：")
        self.trigger_output_k.emit(str(PR_Num_column))#PRN所在列
        self.trigger_output_k.emit( "装备条线名称所在的列数：")
        self.trigger_output_k.emit(str(equip_lines))#装备线所在列
        '''
        temp= "PR FAM所在的列数："+str(PR_FAM_column)+".  \nPR Num所在的列数："+str(PR_Num_column)+'.'
        self.trigger_output_k.emit(temp)
        temp="装备条线所在的列数："+str(equip_lines)
        self.trigger_output_k.emit(temp)
        Num_x_flag=0#标志位：必选是否已经写入过
        Num_o_flag=0#标志位：可选是否已经写入过

        for line in equip_lines:#对每条装备线所在的每一列进行遍历
            PR_Num_list = []#初始化PRN的列表
            PR_FAM_list = []#初始化FAM的列表
            temp=''
            for i in range(1,equip_lines_row+2):
                print(i,str(ws.cell(i, line).value))
                if(str(ws.cell(i, line).value)!='None'):
                    temp=temp+str(ws.cell(i, line).value).replace("\n","-")+'\t'
                #else:
                    #temp+='\t'
            Equipmentline_list.append(temp+'column'+str(line))
            name=str(ws.cell(equip_lines_row+1, line).value)+'-column'+str(line)+'_kopp.txt'
            lines_name=re.sub('[\n/:*?"<>|]','-',name)#去除非法字符
            lines_name.replace("\n","")#去除非法字符
            lines_name.strip(' ')#去除非法字符
            temp="正在计算："+Equipmentline_list[-1]
            self.trigger_output_k.emit(temp)
            print(Equipmentline_list[-1])

            #f = open(lines_name, 'w')#打开文件
            for row in range(start_row, ws_rows_len + 1):#对从头开始到结束的每一行进行遍历
                #print(row,':',str(ws.cell(row, PR_FAM_column).value))
                if ((len(str(ws.cell(row, PR_FAM_column).value).strip(' ').strip('*')))!=3) and ((len(str(ws.cell(row, PR_Num_column).value).strip(' ').strip('*')))!=3) :#是非法行
                    #self.trigger_output_k.emit("警告：PR号列和FAM列都为空，该行被跳过:")
                    #self.trigger_output_k.emit(str(row))#输出跳过非法行
                    continue#跳出循环
                if   str(ws.cell(row, PR_FAM_column).value).strip(' ')=='FAM' or str(ws.cell(row, PR_Num_column).value).strip(' ')=='PRN':#可以跳过
                    continue#跳出循环
                if((len((re.sub('[/:*?"<>|]','-',(str(ws.cell(row, PR_FAM_column).value).strip(' ').strip('*')))))==3) and ((re.sub('[/:*?"<>|]','-',(str(ws.cell(row, PR_FAM_column).value).strip(' ').strip('*'))))).isalpha()):#选择出FAM名称是全字母并且长度为3的行
                            Num_x_flag=1
                            Num_o_flag=1
                            PR_FAM_list.append(str(ws.cell(row, PR_FAM_column).value).strip(' ').strip('*'))
                            if((len(PR_FAM_list)-len(PR_Num_list))==2):
                                PR_Num_list.append('TBD')#补充'TBD'到PRN列表
                                temp="▇▇ 错误 ▇▇：该Family没有对应的PR号："+str(PR_FAM_list[-2])+" ! 程序将\'TBD\' 赋给该Family，如有必要请手动更改."
                                print("错误：出现Family没有对应PR号! 将\'TBD\' 赋给该族:",PR_FAM_list[-2])
                                self.trigger_output_k.emit(temp)
                            if str(ws.cell(row, line).value).strip(' ') == 'X'or str(ws.cell(row, line).value).strip(' ') == 'x' or str(ws.cell(row, line).value).strip(' ') == '●':#必选包对应的字符
                                if ((len(PR_FAM_list) == len(PR_Num_list)) and Num_x_flag == 0):#两个列表相等并且已经必装已经写入过
                                    temp = "▇▇ 错误 ▇▇：同一个Family下有多个必装（X）！行数:" + str(row) + ". FAM名称:" + PR_FAM_list[-1] + "."
                                    self.trigger_output_k.emit(temp)
                                elif ((len(PR_FAM_list) == len(PR_Num_list)) and Num_o_flag == 0 and Option_flag==1):#两个列表相等，选装优先级高，并且已经选装已经写入过
                                    pass#略过必选包
                                else:
                                    PR_Num_list.append(str(ws.cell(row, PR_Num_column).value).strip(' '))#写入必选包
                                    Num_x_flag = 0#将标志位清空
                            elif str(ws.cell(row, line).value).strip(' ') == 'O' or str(ws.cell(row, line).value).strip(' ') == 'o' or str(ws.cell(row, line).value).strip(' ') == '0':#选装包对应的字符
                                if Option_flag==1:
                                    if ((len(PR_FAM_list) == len(PR_Num_list))):
                                        if(Num_o_flag==1):
                                            PR_Num_list[-1] = str(ws.cell(row, PR_Num_column).value).strip(' ')#将最后一位替换为选装包
                                        else:
                                            print("▇▇ 错误 ▇▇：同一个Family下有多个选装（O）！")
                                            temp = "▇▇ 错误 ▇▇：同一个Family下有多个选装（O）！行数:" + str(row) + ". FAM名称:" + PR_FAM_list[-1] + "."
                                            self.trigger_output_k.emit(temp)
                                            #self.trigger_output_k.emit ("错误：同一个Family下有多个选装（O）！行数:")
                                            #self.trigger_output_k.emit (str(row))
                                            #self.trigger_output_k.emit ("FAM名称")
                                            #self.trigger_output_k.emit(PR_FAM_list[-1])  # 输出提示：多个可选包
                                    else:
                                        PR_Num_list.append(str(ws.cell(row, PR_Num_column).value).strip(' '))#直接添加数值到列表最后一位
                                        Num_o_flag=0#标志位清空
                            else:
                                pass
                            continue#跳出循环
                if((len(str(ws.cell(row, PR_Num_column).value).strip(' '))==3) and str(ws.cell(row, PR_Num_column).value).strip(' ').isalnum()):
                    if str(ws.cell(row, line).value).strip(' ') == 'X' or str(ws.cell(row, line).value).strip(' ') == 'x' or str(ws.cell(row, line).value).strip(' ') == '●':  # 必选包对应的字符
                        if ((len(PR_FAM_list) == len(PR_Num_list)) and Num_x_flag == 0):
                            temp = "▇▇ 错误 ▇▇：同一个Family下有多个必装（X）！行数:" + str(row) + ". FAM名称:" + PR_FAM_list[-1] + "."
                            self.trigger_output_k.emit(temp)
                            #self.trigger_output_k.emit("错误：同一个Family下有多个必装（X）！行数:")
                            #self.trigger_output_k.emit(str(row))
                            #self.trigger_output_k.emit("FAM名称")
                            #self.trigger_output_k.emit(PR_FAM_list[-1])  # 输出提示：多个必选包
                        elif ((len(PR_FAM_list) == len(PR_Num_list)) and Num_o_flag == 0 and Option_flag == 1):
                            pass  # 略过必选包
                        else:
                            PR_Num_list.append(str(ws.cell(row, PR_Num_column).value).strip(' '))  # 写入必选包
                            Num_x_flag = 0  # 将标志位清空
                    elif str(ws.cell(row, line).value).strip(' ') == 'O' or str(ws.cell(row, line).value).strip(' ') == 'o' or str(
                            ws.cell(row, line).value).strip(' ') == '0':  # 选装包对应的字符
                        if Option_flag == 1:  # 选装优先级高
                            if ((len(PR_FAM_list) == len(PR_Num_list))):  # 如果两个列表长度相等，说明已经选了自选包或者选装包
                                if (Num_o_flag == 1):  # 如果自选包之前没有写入过
                                    PR_Num_list[-1] = str(ws.cell(row, PR_Num_column).value.strip(' '))  # 将最后一位替换为选装包
                                else:  # 自选包之前已经写入过
                                    print("▇▇ 错误 ▇▇：同一个Family下有多个选装（O）！")
                                    temp = "▇▇ 错误 ▇▇：同一个Family下有多个选装（O）！行数:" + str(row) + ". FAM名称:" + PR_FAM_list[-1] + "."
                                    self.trigger_output_k.emit(temp)
                                    #self.trigger_output_k.emit("错误：同一个Family下有多个选装（O）！行数:")
                                    #self.trigger_output_k.emit(str(row))
                                    #self.trigger_output_k.emit("FAM名称")
                                    #self.trigger_output_k.emit(PR_FAM_list[-1])  # 输出提示：多个可选包
                            else:  # 两个列表长度不等，说明选装包在必选包上面
                                PR_Num_list.append(str(ws.cell(row, PR_Num_column).value).strip(' '))  # 直接添加数值到列表最后一位
                                Num_o_flag = 0  # 标志位清空
                    else:
                            pass
                    continue
            if ((len(PR_FAM_list) - len(PR_Num_list)) == 1):
                PR_Num_list.append('TBD')#加入'TBD'
                temp = "▇▇ 错误 ▇▇：该Family没有对应的PR号：" + str(PR_FAM_list[-1]) + " ! 程序将\'TBD\' 赋给该Family，如有必要请手动更改."
                print("错误：出现Family没有对应PR号! 将\'TBD\' 赋给该族:", PR_FAM_list[-1])
                self.trigger_output_k.emit(temp)
            else:
                pass
            temp="共生成"+str(len(PR_FAM_list))+"个PR FAM和"+str(len(PR_Num_list))+"个PR Num."#输出结果
            self.trigger_output_k.emit(temp)
            dict=zip(PR_FAM_list,PR_Num_list)#生成字典
            '''
            for key,value in dict:#遍历
                f.write(str(key))#写入文件
                f.write(':')
                f.write('\'')
                f.write(str(value))#写入文件
                f.write('\'')
                f.write(',')
            f.close()
            '''
            PR_FAM_list_list.append(PR_FAM_list)
            PR_Num_list_list.append(PR_Num_list)




        #print(PR_FAM_list_list[1])
        if len(PR_FAM_list_list)==0 or len(PR_Num_list_list)==0 or len(Equipmentline_list)==0:
            Error_info = 'kopp文件格式错误，请检查kopp文件'
            return (Error_info, [], [], [])
        Error_info='Error_0'
        print('共计算了',len(PR_Num_list_list),'个装备条线')
        return (Error_info,PR_FAM_list_list,PR_Num_list_list,Equipmentline_list)




