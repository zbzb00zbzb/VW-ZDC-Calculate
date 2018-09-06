from bs4 import BeautifulSoup
import bs4
from openpyxl import Workbook
from openpyxl import load_workbook
import operator
from array import array
import re
import os,os.path
import shutil
import sys
import PyQt5.QtCore
from PyQt5.QtCore import pyqtSignal, QObject
from PyQt5.QtWidgets import (QMainWindow, QTextEdit,
                             QAction, QFileDialog, QApplication, QLabel, QLineEdit, QGridLayout, QWidget, QPushButton,QMessageBox)
from PyQt5.QtGui import QIcon
import winreg

import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException



######################################################################################################################
#path = './nms-PHEV1.4_20170205T170139.html'
#html_file_path ="./A6E-407_____20180410T092938.html"
#status_file_path="C:/Users/SVW/Desktop/5ED_Oktavia NF_FTAB_Status_2018_KW14.xlsx"
#ZDC_file_path="C:\\users\\SVW\\Desktop\\MQB-Octavia NF"
'''标准结构：
<tr>
<td valign="top">
<p class="default_style_b">
<a href="#N68117">0002 - Getriebeelektronik</a>
</p>
</td><td valign="top">
<p class="default_style">AISIN AG6 G4 </p>
</td><td valign="top">
<p class="default_style">09S927158Q </p>
</td><td valign="top">
<p class="default_style">3501</p>
</td><td valign="top">
<p class="default_style">215</p>
</td><td valign="top" style="text-align=center">
<p class="default_style">
<span class="default_ok">0</span>
</p>
</td>
</tr>'''
class html_Analyse_class(QObject):
    trigger_output = PyQt5.QtCore.pyqtSignal(str)
    def __init__(self):
        print('html_Analyse_class chushihua ')

        super().__init__()

        self.initUI()

    def initUI(self):
        self.status_file_path=' '
        self.ZDC_file_path=' '
        self.html_file_path=' '

    def example(self,text):
        print('example')
        self.trigger_output.emit(text)


    def Caculate_all_ZDC_Files(self,status_file_path,ZDC_file_path,html_file_path):

        print('caculate_all_ZDC_Files已启动')
        Error_info='Error_0'
        #self.trigger_output.emit(status_file_path)
        #self.trigger_output.emit(ZDC_file_path)
        #self.trigger_output.emit(html_file_path)

        #trigger_output = PyQt5.QtCore.pyqtSignal(str)
        #html_file_path = "C:/Users/SVW/Desktop/A6E-407_____20180410T092938.html"
        #status_file_path = "C:/Users/SVW/Desktop/5ED_Oktavia NF_FTAB_Status_2018_KW14.xlsx"
        #ZDC_file_path = "C:\\users\\SVW\\Desktop\\MQB-Octavia NF"

        #html_file_path = "C:/Users/EPEI-1/Desktop/NMSNFRPP_17_20180517T104107.html"
        #status_file_path = "C:/Users/EPEI-1/Desktop/ZDC/NMS NF/00-Status/3GB_NMS NF_FTAB_Status_2018_KW17.xlsx"
        #ZDC_file_path = "C:\\Users\\EPEI-1\\Desktop\\ZDC\\NMS NF"

        work_without_diagnose_file = 0#有没有诊断报告文件的标志位
        SG_List_result = []
        desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
        ZDC_File_saved_Path = desktop_path+"\\ProjektZ\\ZDC_Files\\"
        if(os.path.exists(ZDC_File_saved_Path)):
            pass
        else:
            os.makedirs(ZDC_File_saved_Path)
        SG_List_id = [0]
        SG_List_name = [0]
        SG_List_PrNum = [0]
        SG_List_HW_Ver = [0]
        SG_List_SW_Ver = [0]
        SG_List_flag = [0]

        if(len(html_file_path)<4):
            work_without_diagnose_file=1
        else:
            with open(html_file_path, 'r') as f:
                #Soup = BeautifulSoup(f.read(), 'html.parser')
                #Soup = BeautifulSoup(f.read(), 'html.parser', from_encoding='UTF-8')
                Soup = BeautifulSoup(f.read(), 'html.parser', from_encoding='UTF-8')
                Tables = Soup.select('table')
            print("开始分析诊断报告")
            #print(len(Soup.find_all("p", text="SG-Name")))
            if(len(Soup.find_all("p", text="SG-Name")))==0 and (len(Soup.find_all("p", text="SG name")))==0 and (len(Soup.find_all("p", text="控制单元名称")))==0:
                Error_info = '诊断报告不包含控制器零件号列表，请检查诊断报告。'
                return(Error_info,[],[])


            if((len(Soup.find_all("p", text="SG-Name")))>0):
                SG_List_node=Soup.find_all("p", text="SG-Name")[0].parent.parent.parent
            elif len(Soup.find_all("p", text="SG name"))>0:
                SG_List_node=Soup.find_all("p", text="SG name")[0].parent.parent.parent
            else:
                SG_List_node=Soup.find_all("p", text="控制单元名称")[0].parent.parent.parent

            Trs = SG_List_node.select('tr')
            for Tr_n in Trs:#遍历Tr行
                node_p_s = Tr_n.select('p')
                #print('p:', len(node_p_s))
                #print(node_p_s)
                if len(node_p_s)>1:#排除长度不符合标准结构的结点
                    #print('p:',len(node_p_s))
                    node_a_p1 = node_p_s[0].select('a')
                    if len(node_a_p1):
                        if len(str(node_a_p1[0].get_text())) > 6:  # a结点长度要大
                            #print(str(node_a_p1[0].get_text()))
                            SG_List_flag.append('0')
                            SG_List_id.append(str(node_a_p1[0].get_text()).strip(' ').strip('.')[2:4])
                            SG_List_name.append(str(node_a_p1[0].get_text()).strip(' ').strip('.')[7:])
                            SG_List_PrNum.append(str(node_p_s[2].get_text()).strip(' ').strip('.'))
                            SG_List_HW_Ver.append(str(node_p_s[3].get_text()).strip(' ').strip('.'))
                            SG_List_SW_Ver.append(str(node_p_s[4].get_text()).strip(' ').strip('.'))
                        else:
                            break

            SG_List_flag.remove(0)
            SG_List_id.remove(0)
            SG_List_name.remove(0)
            SG_List_PrNum.remove(0)
            SG_List_SW_Ver.remove(0)
            SG_List_HW_Ver.remove(0)
            #print(SG_List_id)
            self.trigger_output.emit("在诊断报告中检测到如下控制器：")
            for SG_List_name_n in SG_List_name:
                print(SG_List_name_n)
                if(len(SG_List_PrNum[SG_List_name.index(SG_List_name_n)])<10):
                    SG_List_PrNum[SG_List_name.index(SG_List_name_n)]+= ' '
                    self.trigger_output.emit(SG_List_id[SG_List_name.index(SG_List_name_n)] + ":\t" + SG_List_PrNum[SG_List_name.index(SG_List_name_n)] + '\t' + SG_List_name_n + ". ")
                    SG_List_PrNum[SG_List_name.index(SG_List_name_n)]= SG_List_PrNum[SG_List_name.index(SG_List_name_n)][:-1]
                else:
                    self.trigger_output.emit(SG_List_id[SG_List_name.index(SG_List_name_n)]+":\t"+SG_List_PrNum[SG_List_name.index(SG_List_name_n)]+'\t'+SG_List_name_n+". ")

            #self.trigger_output.emit("零件号分别为：")
            #for SG_List_PrNum_n in SG_List_PrNum:
                #self.trigger_output.emit(SG_List_PrNum_n)

        if(len(status_file_path)<4):
            Error_info = 'Status文件路径错误'
            return (Error_info, [], [])
        print(status_file_path[-4:])
        if status_file_path[-4:]=='xlsx':
            wb = load_workbook(status_file_path)
            sheets = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheets[0])  # 只导入sheet1页
        elif status_file_path[-4:]=='.xls':
            Book = xlrd.open_workbook(status_file_path)
            print('已打开xls文件')
            index = 0
            nrows, ncols = 0, 0
            sheet = Book.sheet_by_index(0)
            nrows = sheet.nrows
            ncols = sheet.ncols
            # prepare a xlsx sheet
            wb = Workbook()
            sheet1 = wb.get_active_sheet()
            print('xls文件转换为xlsx文件')
            for row in range(0, nrows):
                for col in range(0, ncols):
                    sheet1.cell(row + 1, col + 1).value = str(sheet.cell_value(row, col))
            ws=sheet1

        print("开始分析Status文件")
        ws_rows_len = ws.max_row         #文件行数
        ws_columns_len = ws.max_column    #文件列数
        start_row=2
        start_column=1
        filenames_in_ZDC_Status=[]#存储ZDC文件名的列表
        pass_hide_row=1#标志位，决定是否跳过隐藏行

        XML_column=0
        VBV_Teilenummer_column=0
        FTAB_Teilenummer_column=0
        Diag_Addresse_column = 0
        Ftab_Bezeichnung_column=0

        print('遍历所有列')
        for column in range(start_column, ws_columns_len+1): #遍历所有列
            column_name = str(ws.cell(1, column).value)
            if (column_name=='Softwarepaket') or ('paket' in column_name):
                XML_column = column
            elif(column_name=='Teilenummer in VBV') or (('VBV' in column_name )and ('eilenummer' in column_name)):
                VBV_Teilenummer_column=column
            elif(column_name=='Teilenummer in F-TAB')or (('TAB' in column_name )and ('eilenummer' in column_name)) or (('tab' in column_name )and ('eilenummer' in column_name)):
                FTAB_Teilenummer_column = column
            elif(column_name=='Diag. Addresse') or (('ddresse' in column_name )and ('iag' in column_name))or 'D.A' in column_name:
                Diag_Addresse_column=column
            elif (column_name == 'Ftab Bezeichnung') or ('ezeichnung'in column_name):
                Ftab_Bezeichnung_column=column

        print('------------------------------------',XML_column,VBV_Teilenummer_column,Diag_Addresse_column , Ftab_Bezeichnung_column,ws_rows_len,ws_columns_len)
        if (XML_column == 0):
            Error_info = 'Status文件不包含Softwarepaket列，请检查文件格式。'
            return (Error_info, [], [])
        elif (VBV_Teilenummer_column == 0):
            Error_info = 'Status文件不包含Teilenummer in VBV列，请检查文件格式。'
            return (Error_info, [], [])
        elif (FTAB_Teilenummer_column == 0):
            Error_info = 'Status文件不包含Teilenummer in F-TAB列，请检查文件格式。'
            return (Error_info, [], [])
        elif (Diag_Addresse_column == 0):
            Error_info = 'Status文件不包含Diag. Addresse列，请检查文件格式。'
            return (Error_info, [], [])
        elif (Ftab_Bezeichnung_column == 0) :
            Error_info = 'Status文件不包含Ftab Bezeichnung列，请检查文件格式。'
            return (Error_info, [], [])

        print('遍历所有行')
        if(work_without_diagnose_file==0):
            self.trigger_output.emit("\n在Status文件中根据控制器列表查找对应ZDC文件名：")

        for row in range(start_row, ws_rows_len + 1):  # 遍历所有行
            SG_List_row=[]
            saved_flag = 0
            #PRNR_list = (str(PRNR.childNodes[0].data)).split('+')
            print(row,':',ws.row_dimensions[row].height)
            #if(ws.row_dimensions[row].height == 0) and pass_hide_row:
            if(ws.row_dimensions[row].hidden == True) and pass_hide_row:
                #self.trigger_output.emit("跳过隐藏行:")
                #self.trigger_output.emit(str(row))
                continue#跳出该row
            if ('Swap' in str(ws.cell(row, 1).value)) or str(ws.cell(row, 1).value)=='SWaP-DC' or len(str(ws.cell(row, XML_column).value))<4 or str(ws.cell(row, XML_column).value).strip(' ')[0:2]=='ZZ' :#跳过swap文件
                print('跳过')
                continue  # 跳出该row

            if(work_without_diagnose_file==0):
                #self.trigger_output.emit("在Status文件中根据控制器列表查找对应ZDC文件名：")
                VBV_Teilenummer_str_list =str(ws.cell(row, VBV_Teilenummer_column).value).split('\n')#按照换行符分割
                for VBV_Teilenummer_str in VBV_Teilenummer_str_list:
                    print('-------------------')
                    VBV_Teilenummer_str2 = re.sub('[\n/:*?"<>|]', '-', VBV_Teilenummer_str)  # 去除非法字符
                    VBV_Teilenummer_str3=VBV_Teilenummer_str.replace(" ","").replace(".","").replace("\n", "")   # 去除非法字符
                    print(VBV_Teilenummer_str3)
                    if len(VBV_Teilenummer_str3)<4:
                        continue
                    elif VBV_Teilenummer_str3 in SG_List_PrNum:
                        ########################################3
                        filenames_in_status_list = str(ws.cell(row, XML_column).value).replace('。', '.').replace(' ', '').strip(
                            ' ').split('\n')  # 按照换行符分割
                        SG_List_row.append(str(ws.cell(row, Diag_Addresse_column).value))
                        SG_List_row.append(str(ws.cell(row, Ftab_Bezeichnung_column).value))
                        for filename_in_status in filenames_in_status_list:
                            if len(filename_in_status) < 4:
                                pass
                            else:
                                # print(filename_in_status[-4:])
                                filename_in_status_split_list = filename_in_status.split(' ')  # 按照空格分隔，避免.xml后面还有无用的文字
                                for filename_in_status_split in filename_in_status_split_list:
                                    if len(filename_in_status_split) > 6:
                                        if '.' not in filename_in_status_split:
                                            filename_in_status=filename_in_status+ ".xml"#手动加上后缀
                                            saved_flag = 1
                                            SG_List_row.append(filename_in_status)
                                            if filename_in_status in filenames_in_ZDC_Status:  # 避免重复加入列表
                                                continue
                                            else:
                                                # SG_List_flag[SG_List_PrNum.index(VBV_Teilenummer_str3)]=1#标志位置1,与下面那四句基本等价
                                                for i, SG_List_PrNum_n in enumerate(SG_List_PrNum):
                                                    if SG_List_PrNum_n == VBV_Teilenummer_str3:
                                                        SG_List_flag[i] = 1
                                                filenames_in_ZDC_Status.append(filename_in_status)  # 加入去除xml后面字符功能
                                        elif filename_in_status_split[-4:] != '.XML' and filename_in_status_split[
                                                                                       -4:] != '.xml' and filename_in_status_split[
                                                                                                          -4:] != '.Xml':
                                            temp = '▇▇ 错误 ▇▇：该行定义的数据不是xml文件，请检查第 ' + str(row) + '行：' + str(
                                                filename_in_status_split)
                                            self.trigger_output.emit(temp)
                                            print('▇▇ 错误 ▇▇：该行定义的数据不是xml文件，请检查：', row, ':', filename_in_status_split)

                                        else:
                                            saved_flag = 1
                                            SG_List_row.append(filename_in_status)
                                            if filename_in_status in filenames_in_ZDC_Status:  # 避免重复加入列表
                                                continue
                                            else:
                                                # SG_List_flag[SG_List_PrNum.index(VBV_Teilenummer_str3)]=1#标志位置1,与下面那四句基本等价
                                                for i, SG_List_PrNum_n in enumerate(SG_List_PrNum):
                                                    if SG_List_PrNum_n == VBV_Teilenummer_str3:
                                                        SG_List_flag[i] = 1
                                                filenames_in_ZDC_Status.append(filename_in_status)  # 加入去除xml后面字符功能
                                    else:
                                        pass
                        '''#########################################
                        filenames_in_status_list=str(ws.cell(row, XML_column).value).replace('。','.').strip(' ').split('\n')#按照换行符分割
                        SG_List_row.append(str(ws.cell(row, Diag_Addresse_column).value))
                        SG_List_row.append(str(ws.cell(row, Ftab_Bezeichnung_column).value))
                        for filename_in_status in filenames_in_status_list:
                            saved_flag=1
                            SG_List_row.append(filename_in_status)
                            if filename_in_status in filenames_in_ZDC_Status:#避免重复加入列表
                                continue
                            else:
                                #SG_List_flag[SG_List_PrNum.index(VBV_Teilenummer_str3)]=1#标志位置1,与下面那四句基本等价
                                for i,SG_List_PrNum_n in  enumerate(SG_List_PrNum):
                                    if SG_List_PrNum_n==VBV_Teilenummer_str3:
                                        SG_List_flag[i] = 1
                                filenames_in_ZDC_Status.append(filename_in_status)#加入去除xml后面字符功能
                        ###########################################'''
                    else:
                        continue
                if(saved_flag==0):
                    FTAB_Teilenummer_str_list =str(ws.cell(row, FTAB_Teilenummer_column).value).split('\n')#按照换行符分割
                    for FTAB_Teilenummer_str in FTAB_Teilenummer_str_list:
                        FTAB_Teilenummer_str2 = re.sub('[\n/:*?"<>|]', '-', FTAB_Teilenummer_str)  # 去除非法字符
                        FTAB_Teilenummer_str3=FTAB_Teilenummer_str2.replace(" " ,"").replace(".","").replace("\n", "")  # 去除非法字符
                        print(FTAB_Teilenummer_str3)
                        if FTAB_Teilenummer_str3 in SG_List_PrNum:
                            ######################################
                            filenames_in_status_list = str(ws.cell(row, XML_column).value).replace('。', '.').replace(' ', '').strip(
                                ' ').split('\n')  # 按照换行符分割
                            SG_List_row.append(str(ws.cell(row, Diag_Addresse_column).value))
                            SG_List_row.append(str(ws.cell(row, Ftab_Bezeichnung_column).value))
                            for filename_in_status in filenames_in_status_list:
                                if len(filename_in_status) < 4:
                                    pass
                                else:
                                    # print(filename_in_status[-4:])
                                    filename_in_status_split_list = filename_in_status.split(
                                        ' ')  # 按照空格分隔，避免.xml后面还有无用的文字
                                    for filename_in_status_split in filename_in_status_split_list:
                                        if len(filename_in_status_split) > 6:
                                            if '.' not in filename_in_status_split:  # 不包含"."说明文件缺少xml后缀
                                                filename_in_status = filename_in_status + ".xml"  # 手动加上后缀
                                                saved_flag = 1
                                                SG_List_row.append(filename_in_status)
                                                if filename_in_status in filenames_in_ZDC_Status:  # 避免重复加入列表
                                                    continue
                                                else:
                                                    # SG_List_flag[SG_List_PrNum.index(VBV_Teilenummer_str3)]=1#标志位置1,与下面那四句基本等价
                                                    for i, SG_List_PrNum_n in enumerate(SG_List_PrNum):
                                                        if SG_List_PrNum_n == VBV_Teilenummer_str3:
                                                            SG_List_flag[i] = 1
                                                    filenames_in_ZDC_Status.append(filename_in_status)  # 加入去除xml后面字符功能
                                            elif filename_in_status_split[-4:] != '.XML' and filename_in_status_split[
                                                                                           -4:] != '.xml' and filename_in_status_split[
                                                                                                              -4:] != '.Xml':
                                                temp = '▇▇ 错误 ▇▇：该行定义的数据不是xml文件，请检查：' + str(row) + '行：' + str(
                                                    filename_in_status_split)
                                                self.trigger_output.emit(temp)
                                                print('▇▇ 错误 ▇▇：该行定义的数据不是xml文件，请检查：', row, ':',
                                                      filename_in_status_split)
                                            else:
                                                saved_flag = 1
                                                SG_List_row.append(filename_in_status)
                                                # if filename_in_status.find('SWAP')==0:
                                                if filename_in_status in filenames_in_ZDC_Status:  # 避免重复加入列表
                                                    continue
                                                else:
                                                    #SG_List_flag[SG_List_PrNum.index(VBV_Teilenummer_str3)] = 1  # 标志位置1
                                                    for i, SG_List_PrNum_n in enumerate(SG_List_PrNum):
                                                        if SG_List_PrNum_n == VBV_Teilenummer_str3:
                                                            SG_List_flag[i] = 1
                                                    filenames_in_ZDC_Status.append(filename_in_status)  # 后期可以加入去除xml后面小括号的功能
                                        else:
                                            pass
                            '''######################################
                            filenames_in_status_list = str(ws.cell(row, XML_column).value).replace('。','.').strip(' ').split('\n')#按照换行符分割
                            SG_List_row.append(str(ws.cell(row, Diag_Addresse_column).value))
                            SG_List_row.append(str(ws.cell(row, Ftab_Bezeichnung_column).value))
                            for filename_in_status in filenames_in_status_list:
                                saved_flag=1
                                SG_List_row.append(filename_in_status)
                                #if filename_in_status.find('SWAP')==0:
                                if filename_in_status in filenames_in_ZDC_Status:  # 避免重复加入列表
                                    continue
                                else:
                                    SG_List_flag[SG_List_PrNum.index(VBV_Teilenummer_str3)] = 1  # 标志位置1
                                    filenames_in_ZDC_Status.append(filename_in_status)  # 后期可以加入去除xml后面小括号的功能
                         #####################################  '''
                        else:
                            continue
            else:
                if ('SWaP' in str(ws.cell(row, Ftab_Bezeichnung_column).value)) or str(ws.cell(row, XML_column).value)[0:2] == 'ZZ':  # 跳过swap文件
                    continue  # 跳出该row
                print('程序将保存所有Status文件中定义的ZDC文件',row,str(ws.cell(row, XML_column).value)[0:2])
                if len(str(ws.cell(row, XML_column).value))>4:
                    filenames_in_status_list = str(ws.cell(row, XML_column).value).replace('。','.').replace(' ', '').strip(' ').split('\n')  # 按照换行符分割
                    SG_List_row.append(str(ws.cell(row, Diag_Addresse_column).value))
                    SG_List_row.append(str(ws.cell(row, Ftab_Bezeichnung_column).value))
                    for filename_in_status in filenames_in_status_list:
                        if len(filename_in_status) < 4:
                            pass
                        else:
                            #print(filename_in_status[-4:])
                            filename_in_status_split_list=filename_in_status.split(' ')# 按照空格分隔，避免.xml后面还有无用的文字
                            for filename_in_status_split in filename_in_status_split_list:
                                if len(filename_in_status_split)>6:
                                    if '.' not in filename_in_status_split:  # 不包含"."说明文件缺少xml后缀
                                        print("手动加入后缀xml")
                                        filename_in_status_split = filename_in_status_split + ".xml"  # 手动加上后缀
                                        saved_flag = 1
                                        SG_List_row.append(filename_in_status_split)
                                        filenames_in_ZDC_Status.append(filename_in_status_split)#记录ZDC文件名，用于后面的搜寻
                                        if filename_in_status_split in SG_List_row:  # 避免重复加入列表
                                            continue
                                    elif filename_in_status_split[-4:]!='.XML' and filename_in_status_split[-4:]!='.xml'and filename_in_status_split[-4:]!='.Xml':
                                        temp='▇▇ 错误 ▇▇：该行定义的数据不是xml文件，请检查：'+str(row)+'行：'+str(filename_in_status_split)
                                        self.trigger_output.emit(temp)
                                        print('▇▇ 错误 ▇▇：该行定义的数据不是xml文件，请检查：',row,':',filename_in_status_split)

                                    else:
                                        saved_flag = 1
                                        SG_List_row.append(filename_in_status_split)
                                        filenames_in_ZDC_Status.append(filename_in_status_split)#记录ZDC文件名，用于后面的搜寻
                                        if filename_in_status_split in SG_List_row:  # 避免重复加入列表
                                            continue
                                else:
                                    pass
            if (saved_flag == 1):
                SG_List_result.append(SG_List_row)

        print(SG_List_result)
        print(SG_List_flag)

        if(work_without_diagnose_file==0):
            #print('filenames_in_ZDC_Status:',filenames_in_ZDC_Status)
            #print('SG_List_flag',SG_List_flag)
            SG_not_exist_list=[i for i, x in enumerate(SG_List_flag) if x == '0']
            for SG_not_exist_n in SG_not_exist_list:
                print("该控制器未找到对应ZDC：",SG_List_name[SG_not_exist_n])
                self.trigger_output.emit("▇ 注意 ▇ 该控制器: "+SG_List_id[SG_not_exist_n]+"\t" +SG_List_name[SG_not_exist_n]+'\t'+SG_List_PrNum[SG_not_exist_n]+" 在Status表中没有对应的ZDC, 如果需要请手动添加. ")
        else:
            print('未加载诊断报告，因此保存Status文件中所有控制器对应的ZDC文件')
            self.trigger_output.emit("未加载诊断报告，因此程序将保存Status文件中所有控制器对应的所有ZDC文件.")

        #key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\Explorer\ShellFolders' )
        #desktop_path= winreg.QueryValueEx(desktop_key, "Desktop")[0]
        #print(desktop_path)
        print('开始遍历')
        if(len(ZDC_file_path)<4):
            Error_info = 'ZDC文件所在路径错误，请检查。'
            return (Error_info, [], [])
        for root, subdirs, files in os.walk(ZDC_file_path):
            for filepath in files:#遍历
                #print(filepath)
                if filepath in filenames_in_ZDC_Status:#如果ZDC列表中存在
                   print('已找到:',filepath,'.')
                   filenames_in_ZDC_Status[filenames_in_ZDC_Status.index(filepath)]='0'#已找到的文件进行标记
                   oldfile =os.path.join(root, filepath)#老文件地址
                   #newfile = desktop_path+"\\PyQt5" + filepath#新文件地址
                   newfile = ZDC_File_saved_Path+ filepath#新文件地址
                   print(oldfile)
                   print(newfile)
                   shutil.copyfile(oldfile, newfile)#移动文件到新目录

        for file in filenames_in_ZDC_Status:
            if(file !='0' and len(str(file))>6):
                #self.trigger_output.emit("▇ 注意 ▇ 该ZDC文件未找到："+file+", \t\t如果需要请手动添加.")
                for SG_List_result_n in SG_List_result:
                    for SG_List_result_n_i in SG_List_result_n[2:]:
                        if str(file)==str(SG_List_result_n_i):
                            SG_List_result[SG_List_result.index(SG_List_result_n)].remove(SG_List_result_n_i)
                            self.trigger_output.emit("▇ 注意 ▇ 该控制器: "+SG_List_result[SG_List_result.index(SG_List_result_n)][0]+"  "+SG_List_result[SG_List_result.index(SG_List_result_n)][1]+"  对应的ZDC文件: "+file+"不在指定的文件夹中,\t如果需要请手动添加.")

        print(SG_List_result)
        #return filenames_in_ZDC_Status
        print('end')
        Error_info='Error_0'
        return Error_info,SG_List_result,ZDC_File_saved_Path