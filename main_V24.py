import sys
import PyQt5
from PyQt5.QtWidgets import (QMainWindow, QTextEdit,QTableWidget,QTableWidgetItem,QComboBox,QCheckBox,
                             QAction, QFileDialog, QApplication, QLabel, QLineEdit, QGridLayout, QWidget, QPushButton,QMessageBox,QTextBrowser,QPlainTextEdit, QTabWidget, QVBoxLayout)
import ctypes
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import shutil
import os,os.path
import xml.dom.minidom
from pyDes import *
import base64
import win32api
from xml.dom.minidom import parse
from html_analyse_V12 import html_Analyse_class
from ZDC_Xml_analyse_v12 import ZDC_file_Calculate_class
from dipon2txt_v7 import Kopp_Analyse_and_ZDC_Calculate_class
from Macro_generate import Macro_generate_class
from openpyxl import Workbook
from openpyxl import load_workbook
import xlrd
import xlwt


class ZDC_file_locate_Widget(QWidget):
    trigger_output_main = PyQt5.QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__()
        print('ZDC_file_locate_Widget: chushihua ')
        self.initUI()

    def initUI(self):
        self.html_Analyse_instance=html_Analyse_class()#初始 实例化
        self.html_Analyse_instance.trigger_output.connect(self.display_html_Analyse_output)


        self.Error_info_1=''
        self.Error_info_2=''
        self.Error_info_3=''
        self.SG_List_result=[]
        self.checkbox_SG_list=[]
        self.ZDC_status_str=' '
        self.ZDC_file_path_str=' '
        self.Diagnose_report_str=' '
        desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
        self.ZDC_File_saved_Path =  desktop_path+"\\ProjektZ\\ZDC_Files\\"
        #self.ZDC_File_saved_Path = "C:\\users\\SVW\\Desktop\\ZDC_Files\\"
        if(os.path.exists(self.ZDC_File_saved_Path)):
            pass
        else:
            os.makedirs(self.ZDC_File_saved_Path)


        self.ZDC_status = QLabel('Status文件')
        self.ZDC_file_path = QLabel('ZDC文件地址')
        self.Diagnose_report = QLabel('诊断报告')

        self.ZDC_status_Edit = QLineEdit(self)
        self.ZDC_file_path_Edit = QLineEdit(self)
        self.Diagnose_report_Edit = QLineEdit(self)

        self.grid2 = QGridLayout()
        grid=self.grid2
        grid.setSpacing(16)

        grid.addWidget(self.ZDC_status, 1, 1,1,1)
        grid.addWidget(self.ZDC_status_Edit, 1, 2,1,4)
        qbtn_ZDC_status = QPushButton('…', self)
        qbtn_ZDC_status.clicked.connect(self.get_ZDC_status_Filename)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(qbtn_ZDC_status, 1, 6,1,1)

        grid.addWidget(self.ZDC_file_path, 2, 1,1,1)
        grid.addWidget(self.ZDC_file_path_Edit, 2, 2,1,4)
        qbtn_ZDC_file_path = QPushButton('…', self)
        qbtn_ZDC_file_path.clicked.connect(self.get_ZDC_file_path)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(qbtn_ZDC_file_path, 2, 6,1,1)

        grid.addWidget(self.Diagnose_report, 3, 1,1,1)
        grid.addWidget(self.Diagnose_report_Edit, 3, 2,1,4)
        qbtn_Diagnose_report = QPushButton('…', self)
        qbtn_Diagnose_report.clicked.connect(self.get_Diagnose_report_Filename)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(qbtn_Diagnose_report, 3, 6,1,1)

        qbtn_ZDC_find_calculate = QPushButton('自动生成ZDC列表', self)
        qbtn_ZDC_find_calculate.clicked.connect(self.Caculate_all_ZDC_Filenames)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(qbtn_ZDC_find_calculate, 5,5, 1, 2)

        #self.ZDC_find_calculate_output= QTextBrowser(self)
        #grid.addWidget(self.ZDC_find_calculate_output, 16,16,1,1)

        self.SG_id_and_ZDC_Table = QTableWidget(50,3)
        self.SG_id_and_ZDC_Table.setHorizontalHeaderLabels(['ID','描述','ZDC文件名'])
        self.SG_id_and_ZDC_Table.verticalHeader().setHidden(True);
        grid.addWidget(self.SG_id_and_ZDC_Table,6,1,13,6)
        self.SG_id_and_ZDC_Table.setColumnWidth(0, 50)
        self.SG_id_and_ZDC_Table.setColumnWidth(1, 150)
        self.SG_id_and_ZDC_Table.setColumnWidth(2, 400)
        self.SG_id_and_ZDC_Table.setColumnWidth(3, 50)

        self.qbtn_ZDC_add_file = QPushButton('手动添加ZDC文件', self)
        self.qbtn_ZDC_add_file.clicked.connect(self.Mannuel_add_ZDC_file_Funktion)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(self.qbtn_ZDC_add_file, 5, 1, 1, 2)


        self.Kopp_Analyse_and_ZDC_Calculate_instance = Kopp_Analyse_and_ZDC_Calculate_class()  # 初始 实例化
        self.Kopp_Analyse_and_ZDC_Calculate_instance.trigger_output_k.connect(
            self.display_Kopp_Analyse_and_ZDC_Calculate_output)


        self.line_index=0
        self.result = [0]
        self.PR_FAM_list_list = []
        self.PR_Num_list_list = []
        self.equipment_lines = []
        self.option_flag = 1
        self.kopp_file_path_str = ' '
        self.ZDC_file_list = [0]
        self.equipment_lines_list_exist_index = 1
        self.PR_change_Automatically_index=0

        self.kopp_file_path = QLabel('Kopp/Dipon文件')
        self.kopp_file_path_Edit = QLineEdit(self)


        grid.addWidget(self.kopp_file_path, 1, 8,1,1)
        grid.addWidget(self.kopp_file_path_Edit, 1,9, 1,5)
        qbtn_kopp_file_path = QPushButton('…', self)
        qbtn_kopp_file_path.clicked.connect(self.get_kopp_file_path)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(qbtn_kopp_file_path, 1, 14,1,1)

        self.option_flag_checkbox = QCheckBox('不添加选装包', self)
        self.option_flag_checkbox.stateChanged.connect(self.option_flag_checkbox_changed)
        grid.addWidget(self.option_flag_checkbox, 1, 15,1,1)
        #self.option_flag_checkbox.state = PyQt5.QtCore.Qt.Checked

        self.qbtn_Caculate_kopp_and_ZDC_Files = QPushButton('解析\nKopp\n文件', self)
        self.qbtn_Caculate_kopp_and_ZDC_Files.clicked.connect(self.Caculate_kopp_and_ZDC_Files)
        qbtn_ZDC_status.resize(self.qbtn_Caculate_kopp_and_ZDC_Files.sizeHint())
        grid.addWidget(self.qbtn_Caculate_kopp_and_ZDC_Files, 1,16,3,1)

        self.qbtn_Caculate_kopp_and_ZDC_Files = QPushButton('导入PR号', self)
        self.qbtn_Caculate_kopp_and_ZDC_Files.clicked.connect(self.import_PR_Nums)
        qbtn_ZDC_status.resize(self.qbtn_Caculate_kopp_and_ZDC_Files.sizeHint())
        grid.addWidget(self.qbtn_Caculate_kopp_and_ZDC_Files, 2,15,1,1)

        self.qbtn_Caculate_kopp_and_ZDC_Files = QPushButton('导出PR号', self)
        self.qbtn_Caculate_kopp_and_ZDC_Files.clicked.connect(self.export_PR_Nums)
        qbtn_ZDC_status.resize(self.qbtn_Caculate_kopp_and_ZDC_Files.sizeHint())
        grid.addWidget(self.qbtn_Caculate_kopp_and_ZDC_Files, 3,15,1,1)

        #######################################################################################
        self.ZDC_file_path = QLabel('输出信息栏')
        grid.addWidget(self.ZDC_file_path, 14,7,1,9)

        self.Caculate_kopp_and_ZDC_Files_output = QTextBrowser(self)
        grid.addWidget(self.Caculate_kopp_and_ZDC_Files_output, 15,7,4,9)

        self.PR_FAM_and_Num_Table = QTableWidget(2, 200)
        self.PR_FAM_and_Num_Table.setVerticalHeaderLabels(['PR FAM', 'PR Num'])
        self.PR_FAM_and_Num_Table.horizontalHeader().setHidden(True);
        grid.addWidget(self.PR_FAM_and_Num_Table, 2,8,2,7)
        self.PR_FAM_and_Num_Table.itemChanged.connect(self.update_PR_FAM_or_Num)
        #self.PR_FAM_and_Num_Table.currentItemChanged.connect(self.update_PR_FAM_or_Num)

        self.equipment_lines_label = QLabel('装备条线选择：')
        grid.addWidget(self.equipment_lines_label, 5,8,1,1)

        self.equipment_lines_list = QComboBox(self)
        grid.addWidget(self.equipment_lines_list, 5,9,1,5)

        #'''
        self.PR_Num_Lookup_input = QLineEdit(self)
        grid.addWidget(self.PR_Num_Lookup_input, 5,14,1,2)

        self.qtbn_PR_Num_Lookup_input = QPushButton('查找', self)
        self.qtbn_PR_Num_Lookup_input.clicked.connect(self.lookup_PR_Num_Funktion)
        grid.addWidget(self.qtbn_PR_Num_Lookup_input,5,16,1,1)

        self.ZDC_file_Calculate_instance = ZDC_file_Calculate_class()  # 初始 实例化
        self.ZDC_file_Calculate_instance.trigger_output_Z.connect(self.display_ZDC_file_Calculate_instance_output)
        self.Macro_generate_instance=Macro_generate_class()#初始化
        self.Macro_generate_instance.trigger_output_M.connect(self.display_Macro_generate_instance_output)



        #self.PR_Num_list = []
        self.ZDC_file_list = []#ZDC文件的list
        self.ZDC_data_list = []#每个ZDC文件的数据
        self.All_controller_data_and_programming=[]


        self.qbtn_ZDC_file_Calculate = QPushButton('计算\nZDC\n数据', self)
        self.qbtn_ZDC_file_Calculate.clicked.connect(self.ZDC_file_Calculate_funktion)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(self.qbtn_ZDC_file_Calculate, 7, 16, 2, 1)

        self.qbtn_ZDC_file_Calculate = QPushButton('保存\nZDC\n数据', self)
        self.qbtn_ZDC_file_Calculate.clicked.connect(self.Save_ZDC_data_funktion)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(self.qbtn_ZDC_file_Calculate, 9, 16, 2, 1)

        self.qbtn_macro_generate_Widget = QPushButton('生成\nMacro\n文件', self)
        self.qbtn_macro_generate_Widget.clicked.connect(self.macro_generate_funktion)
        # qbtn_ZDC_status.resize(qbtn_ZDC_status.sizeHint())
        grid.addWidget(self.qbtn_macro_generate_Widget, 11, 16, 2, 1)

        #self.ZDC_data_output_TextBrowser = QTextBrowser(self)
        #grid.addWidget(self.ZDC_data_output_TextBrowser, 12,6,4,9)

        self.ZDC_data_output_Table = QTableWidget(200, 4)
        self.ZDC_data_output_Table.setHorizontalHeaderLabels(['ID', '数据地址','数据内容',' '])
        self.ZDC_data_output_Table.verticalHeader().setHidden(True);
        self.ZDC_data_output_Table.setColumnWidth(0, 50)
        self.ZDC_data_output_Table.setColumnWidth(1, 100)
        self.ZDC_data_output_Table.setColumnWidth(2, 1000)
        # self.ZDC_data_output_Table.setVerticalHeaderLabels(['PR FAM','PR Num'])
        grid.addWidget(self.ZDC_data_output_Table, 7,7,7,9)

        self.setLayout(grid)
        self.setGeometry(100, 100, 550, 500)
        self.setWindowTitle('ZDC_file_locate_Widget')
        self.show()
    def  get_ZDC_status_Filename(self):
        fname = QFileDialog.getOpenFileName(self, 'Open file','.',"xlsx files (*.xlsx *.xls)")
        self.ZDC_status_str=fname
        print(self.ZDC_status_str)
        self.flash_strs()
        self.ZDC_status_Edit.setText(self.ZDC_status_str[0])

    def  get_ZDC_file_path(self):
        fname = QFileDialog.getExistingDirectory(self, '/home')
        self.ZDC_file_path_str=fname
        print(self.ZDC_file_path_str)
        self.flash_strs()
        self.ZDC_file_path_Edit.setText(self.ZDC_file_path_str)

    def  get_Diagnose_report_Filename(self):
        fname = QFileDialog.getOpenFileName(self, 'Open file', '.',"html files (*.html)")
        self.Diagnose_report_str=fname
        print(self.Diagnose_report_str)
        self.flash_strs()
        self.Diagnose_report_Edit.setText(self.Diagnose_report_str[0])


    def flash_strs(self):
        print("flashing")
        self.ZDC_status_Edit.setText(self.ZDC_status_str[0])
        self.ZDC_file_path_Edit.setText(self.ZDC_file_path_str)
        self.Diagnose_report_Edit.setText(self.Diagnose_report_str[0])
        #print(self.ZDC_file_path_str)

    def Caculate_all_ZDC_Filenames(self):
        self.SG_id_and_ZDC_Table.clear()
        self.SG_id_and_ZDC_Table.setHorizontalHeaderLabels(['ID','描述','ZDC文件名'])
        self.Caculate_kopp_and_ZDC_Files_output.clear()
        print('Caculate_all_ZDC_Filenames')
        print(self.ZDC_file_path_str)
        print(self.ZDC_status_str[0])
        print(self.Diagnose_report_str[0])
        text = "777"
        #self.html_Analyse_instance.caculateallZDCFiles(text)
        self.Error_info_1,self.SG_List_result,self.ZDC_File_saved_Path= self.html_Analyse_instance.Caculate_all_ZDC_Files(self.ZDC_status_str[0], self.ZDC_file_path_str, self.Diagnose_report_str[0])
        #self.html_Analyse_instance.example(text)
        if self.Error_info_1=='Error_0':
            count=0
            if(len(self.SG_List_result)>0):

                if len(self.SG_List_result) > self.SG_id_and_ZDC_Table.rowCount():
                    self.SG_id_and_ZDC_Table.setRowCount(len(self.SG_List_result) + 30)  # 需要统计SG_List_result列表中第三个元素的总和作为总行数，后期再优化

                for index,SG_List_result_n in enumerate(self.SG_List_result):
                    SG_id_Item = QTableWidgetItem(str(SG_List_result_n[0]))
                    SG_besch_Item = QTableWidgetItem(str(SG_List_result_n[1]))

                    if (index+len(SG_List_result_n)) > self.SG_id_and_ZDC_Table.rowCount():
                        self.SG_id_and_ZDC_Table.setRowCount((index+len(SG_List_result_n)) + 10)  # 需要统计SG_List_result列表中第三个元素的总和作为总行数，后期再优化

                    self.SG_id_and_ZDC_Table.setItem(index+count, 0, SG_id_Item)
                    self.SG_id_and_ZDC_Table.setItem(index+count, 1, SG_besch_Item)

                    for i in range(2, (len(SG_List_result_n))):
                        SG_ZDC_Item = QTableWidgetItem(str(SG_List_result_n[i]))
                        checkBox = QTableWidgetItem(SG_ZDC_Item);
                        #checkBox = QTableWidgetItem(' ');
                        checkBox.setCheckState(Qt.Checked);
                        #checkbox_SG_list.append(checkBox)

                        if (index + count+i-1) > self.SG_id_and_ZDC_Table.rowCount():
                            self.SG_id_and_ZDC_Table.setRowCount((index + count+i-1) + 10)  # 需要统计SG_List_result列表中第三个元素的总和作为总行数，后期再优化

                        self.SG_id_and_ZDC_Table.setItem(index+count+i-2, 2, checkBox)
                        #self.SG_id_and_ZDC_Table.setItem(index+count+i-2, 3, SG_ZDC_Item)
                    count=count+len(SG_List_result_n)-3
        elif self.Error_info_1=='Error_1':
            QMessageBox.information(self, "错误", "诊断报告有问题")
        elif self.Error_info_1 == 'Error_2':
            QMessageBox.information(self, "错误", "ZDC路径有问题")
        elif self.Error_info_1 == 'Error_3':
            QMessageBox.information(self, "错误", "ZDC Status文件有问题")
        else:
            QMessageBox.information(self, "错误", self.Error_info_1)


    def Mannuel_add_ZDC_file_Funktion(self):
        filepath_and_filename = QFileDialog.getOpenFileName(self, 'Open file', '.', "xml files (*.xml)")
        if(len(filepath_and_filename[0])>8):
            dom = xml.dom.minidom.parse(filepath_and_filename[0])
            print('已成功解析文件')
            ZDC_root = dom.documentElement
            VORSCHRIFT = ZDC_root.getElementsByTagName("VORSCHRIFT")
            if len(VORSCHRIFT)>0:
                DIREKT = VORSCHRIFT[0].getElementsByTagName("DIREKT")
                DIAGN = DIREKT[0].getElementsByTagName("DIAGN")
                ADR = DIAGN[0].getElementsByTagName("ADR")
                Controller_id = str(ADR[0].childNodes[0].data)
            else:
                QMessageBox.information(self, "错误", "ZDC文件格式错误，请检查。")
                return
            filepath_and_filename_list=filepath_and_filename[0].split('/')
            filename=filepath_and_filename_list[-1]
            print(filename)
            new_filepath_and_name = self.ZDC_File_saved_Path + filename  # 新文件地址
            print(filepath_and_filename[0].replace('/','\\'))
            print(new_filepath_and_name)
            shutil.copyfile(filepath_and_filename[0].replace('/','\\'), new_filepath_and_name)
            print('io')
            self.SG_List_result.append([Controller_id,'  ',filename])
            count=0

            if len(self.SG_List_result) >= self.SG_id_and_ZDC_Table.rowCount():
                self.SG_id_and_ZDC_Table.setRowCount(
                    len(self.SG_List_result) + 10)  # 需要统计SG_List_result列表中第三个元素的总和作为总行数，后期再优化

            for index, SG_List_result_n in enumerate(self.SG_List_result):
                SG_id_Item = QTableWidgetItem(str(SG_List_result_n[0]))
                SG_besch_Item = QTableWidgetItem(str(SG_List_result_n[1]))
                self.SG_id_and_ZDC_Table.setItem(index + count, 0, SG_id_Item)
                self.SG_id_and_ZDC_Table.setItem(index + count, 1, SG_besch_Item)

                for i in range(2, (len(SG_List_result_n))):
                    SG_ZDC_Item = QTableWidgetItem(str(SG_List_result_n[i]))
                    checkBox = QTableWidgetItem(SG_ZDC_Item);
                    # checkBox = QTableWidgetItem(' ');
                    checkBox.setCheckState(Qt.Checked);
                    # checkbox_SG_list.append(checkBox)
                    self.SG_id_and_ZDC_Table.setItem(index + count + i - 2, 2, checkBox)
                    # self.SG_id_and_ZDC_Table.setItem(index+count+i-2, 3, SG_ZDC_Item)

                count = count + len(SG_List_result_n) - 3

    def display_html_Analyse_output(self,message):
        #print('mesage:',message)
        #self.Diagnose_report_Edit.setText(message)
        self.Caculate_kopp_and_ZDC_Files_output.append(message)


    def get_kopp_file_path(self):
        fname = QFileDialog.getOpenFileName(self, 'Open file', '.', "xlsx files (*.xlsx)")
        self.kopp_file_path_str = fname
        print(self.kopp_file_path_str)
        self.flash_strs()

    def flash_strs(self):
        self.kopp_file_path_Edit.setText(self.kopp_file_path_str[0])

    # def return_PR_Num(self):
    # self.return_PR_Num_pyqtSignal.emit(self.PR_Num_list_list[self.line_index])

    def import_PR_Nums(self):
        print("import_PR_Nums")
        import_file_path = QFileDialog.getOpenFileName(self, 'Open file','.',"xlsx files (*.xlsx *.xls)")
        print(import_file_path)
        status_file_path=import_file_path[0]
        if (len(status_file_path) < 4):
            QMessageBox.information(self, "错误", "文件路径错误！")
            return
        print(status_file_path)
        if status_file_path[-4:] == 'xlsx':
            wb = load_workbook(status_file_path)
            sheets = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheets[0])  # 只导入sheet1页
            start_row = 3

        elif status_file_path[-4:] == '.xls':
            Book = xlrd.open_workbook(status_file_path)
            print('已打开xls文件')
            index = 0
            nrows, ncols = 0, 0
            sheet = Book.sheet_by_index(0)
            nrows = sheet.nrows
            ncols = sheet.ncols
            # prepare a xlsx sheet
            wb = Workbook()
            sheet1 = wb.get_active_sheet()
            print('xls文件转换为xlsx文件')
            for row in range(0, nrows):
                for col in range(0, ncols):
                    sheet1.cell(row + 1, col + 1).value = str(sheet.cell_value(row, col))
            ws = sheet1
            start_row = 2


        ws_rows_len = ws.max_row         #文件行数
        ws_columns_len = ws.max_column    #文件列数
        import_PRN = []
        import_FAM = []
        error_flag = 0

        for row in range(start_row, ws_rows_len + 1):  # 遍历所有行

            if len(str(ws.cell(row, 1).value))==3:
                import_PRN.append(str(ws.cell(row, 1).value))
            else:
                import_PRN.append(str(ws.cell(row, 1).value))
                error_flag=1

            if len(str(ws.cell(row, 3).value))==3:
                import_FAM.append(str(ws.cell(row, 3).value))
            else:
                import_FAM.append(str(ws.cell(row, 3).value))
                error_flag=1


        self.PR_FAM_list_list.append(import_FAM)
        self.PR_Num_list_list.append(import_PRN)
        print(self.PR_Num_list_list)
        import_equipment_lines=[str(status_file_path)]
        self.equipment_lines_list.addItems(import_equipment_lines)


        self.line_index=len(self.PR_Num_list_list)-1
        print('self.line_index',self.line_index)
        self.equipment_lines_list.setCurrentIndex(self.line_index)
        self.Display_current_PR_combination(self.line_index)
        print('555')
        if error_flag==1:
            info="已导入"+str(len(import_PRN))+"个PR-Nr和Familie，但是存在格式不标准的情况，请检查源文件。"
        else:
            info="已导入"+str(len(import_PRN))+"个PR-Nr和Familie。"
        QMessageBox.information(self, "通知", info)
        print('end')

    def export_PR_Nums(self):
        print("export_PR_Nums")
        if len(self.PR_Num_list_list) < 1:
            QMessageBox.information(self, "错误", "请先导入kopp文件，再导出PR号。")
        else:
            desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
            saved_file_path = QFileDialog.getSaveFileName(self, 'save file', desktop_path,
                                                          "Excel files (*.xls);;all files(*.*)")
            if (len(saved_file_path[0])) < 6:
                return 0
            print('222')
            workbook = xlwt.Workbook(encoding='ascii')
            pattern = xlwt.Pattern()  # Create the Pattern
            pattern.pattern = xlwt.Pattern.SOLID_PATTERN  # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
            pattern.pattern_fore_colour = 5  # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
            style = xlwt.XFStyle()  # Create the Pattern
            style.pattern = pattern  # Add Pattern to Style
            print('333')
            pattern2 = xlwt.Pattern()  # Create the Pattern
            pattern2.pattern = xlwt.Pattern.SOLID_PATTERN  # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
            pattern2.pattern_fore_colour = 2  # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
            style2 = xlwt.XFStyle()  # Create the Pattern
            style2.pattern = pattern2  # Add Pattern to Style

            worksheet = workbook.add_sheet('PR sheet')
            row_count = 1
            worksheet.write(0, 0, 'PR-Nr')
            worksheet.write(0, 2, 'Familie')
            print(self.PR_FAM_list_list)
            print(self.line_index)
            print(len(self.PR_FAM_list_list[self.line_index]))
            for count_n in range(0,len(self.PR_FAM_list_list[self.line_index])):
                print(count_n)
                worksheet.write(row_count, 0, self.PR_Num_list_list[self.line_index][count_n])
                worksheet.write(row_count, 2, self.PR_FAM_list_list[self.line_index][count_n])
                row_count += 1

            print('335')

            workbook.save(saved_file_path[0])
            QMessageBox.information(self, "通知", "PR-Nr和Familie已导出。")

    def Caculate_kopp_and_ZDC_Files(self):
        self.equipment_lines_list.clear()
        self.Caculate_kopp_and_ZDC_Files_output.clear()
        print('Caculate_all_ZDC_Filenames')
        #self.PR_Num_list_list=[]
        #print(len( self.PR_Num_list_list))
        # self.html_Analyse_instance.caculateallZDCFiles(text)
        self.Error_info_2,self.PR_FAM_list_list, self.PR_Num_list_list, self.equipment_lines = self.Kopp_Analyse_and_ZDC_Calculate_instance.Kopp_Analyse_and_ZDC_Calculate(
            self.kopp_file_path_str[0], self.option_flag)
        # self.html_Analyse_instance.example(text)
        # print(self.result)
        self.PR_change_Automatically_index =1

        if self.Error_info_2 == 'Error_0':
            #self.PR_change_Automatically_index = 1
            PR_FAM_list = self.PR_FAM_list_list[1]
            PR_Num_list = self.PR_Num_list_list[1]
            for i in PR_FAM_list:
                # print(i,PR_FAM_list.index(i),str(PR_Num_list[PR_FAM_list.index(i)]))
                #print('PR_change_Automatically_index',self.PR_change_Automatically_index,'str(i)',str(i))
                FAM_Item = QTableWidgetItem(str(i))
                Num_Item = QTableWidgetItem(str(PR_Num_list[PR_FAM_list.index(i)]))
                self.PR_FAM_and_Num_Table.setItem(0, PR_FAM_list.index(i), FAM_Item)
                self.PR_FAM_and_Num_Table.setItem(1, PR_FAM_list.index(i), Num_Item)
                self.PR_FAM_and_Num_Table.setColumnWidth(PR_FAM_list.index(i), 50)
            self.equipment_lines_list.addItems(self.equipment_lines)
        else:
            QMessageBox.information(self,"错误",self.Error_info_2)
        print("EOL")
        if self.equipment_lines_list_exist_index:
            self.equipment_lines_list.currentIndexChanged.connect(self.Display_current_PR_combination)
            self.equipment_lines_list_exist_index = 0
        self.PR_change_Automatically_index =0

    def Display_current_PR_combination(self, line_index):
        self.PR_change_Automatically_index=1
        self.PR_FAM_and_Num_Table.clear()
        self.PR_FAM_and_Num_Table.setVerticalHeaderLabels(['PR FAM', 'PR Num'])
        self.PR_FAM_and_Num_Table.horizontalHeader().setHidden(True);
        self.line_index=line_index
        print(self.line_index)
        if (len(self.equipment_lines_list) > 0):
            print("xianshi")
            PR_FAM_list = self.PR_FAM_list_list[line_index]
            PR_Num_list = self.PR_Num_list_list[line_index]
            print(PR_FAM_list)
            print(PR_Num_list)
            for i in PR_FAM_list:
                # print(i,PR_FAM_list.index(i),str(PR_Num_list[PR_FAM_list.index(i)]))
                FAM_Item = QTableWidgetItem(str(i))
                Num_Item = QTableWidgetItem(str(PR_Num_list[PR_FAM_list.index(i)]))
                self.PR_FAM_and_Num_Table.setItem(0, PR_FAM_list.index(i), FAM_Item)
                self.PR_FAM_and_Num_Table.setItem(1, PR_FAM_list.index(i), Num_Item)
                self.PR_FAM_and_Num_Table.setColumnWidth(PR_FAM_list.index(i), 50)
            #QMessageBox.information(self, "PR", "PR号已更新")
        self.PR_change_Automatically_index =0

    def display_Kopp_Analyse_and_ZDC_Calculate_output(self, message):
        #print('mesage:', message)
        # self.Diagnose_report_Edit.setText(message)
        self.Caculate_kopp_and_ZDC_Files_output.append(message)

    def option_flag_checkbox_changed(self, state):
        if state == PyQt5.QtCore.Qt.Checked:
            self.option_flag = 0
        else:
            self.option_flag = 1
        # print(self.equipment_lines_list.count())
        if self.equipment_lines_list.count() > 0:
            QMessageBox.information(self, "PR", "请重新点击解析kopp文件按钮。")


    def lookup_PR_Num_Funktion(self):
        PR_Num = self.PR_Num_Lookup_input.text().upper()
        if len(self.PR_FAM_list_list)>0:
            if(len(PR_Num)!=3):
                QMessageBox.information(self, "PR", "请输入格式正确的PR号。")
            else:
                #print(self.PR_Num_list_list[self.line_index])
                #print(self.PR_FAM_list_list[self.line_index])
                if(PR_Num in self.PR_Num_list_list[self.line_index]):
                    print('Num')
                    item = self.PR_FAM_and_Num_Table.findItems(PR_Num, Qt.MatchExactly)#遍历
                    #print(item)
                    row = item[0].row()
                    column=item[0].column()
                    #print("DDD", row,column)#获取其行号
                    self.PR_FAM_and_Num_Table.horizontalScrollBar().setSliderPosition(column)
                    #print(self.PR_FAM_and_Num_Table.findItems(PR_Num))
                elif PR_Num in self.PR_FAM_list_list[self.line_index]:
                    print('FAM')
                    item = self.PR_FAM_and_Num_Table.findItems(PR_Num, Qt.MatchExactly)#遍历
                    #print(item)
                    row = item[0].row()
                    column=item[0].column()
                    #print("DDD", row,column)#获取其行号
                    self.PR_FAM_and_Num_Table.horizontalScrollBar().setSliderPosition(column)  #定位
                else:
                    QMessageBox.information(self, "错误", "该PR号不存在。")
    def update_PR_FAM_or_Num(self,item):
        if(self.PR_change_Automatically_index==1):
            pass
        elif(len(item.text())!=3):
            QMessageBox.information(self, "错误", "该输入格式正确的PR号。")
            return
            #self.Display_current_PR_combination(self.line_index)
        else:
            if(len(self.PR_Num_list_list)>0 and len(self.PR_FAM_list_list)>0):
                row=item.row()
                column=item.column()
                print(row,column,item.text().upper())
                print('len',len(self.PR_Num_list_list) ,len(self.PR_FAM_list_list))
                if(row==0):
                    if(len(self.PR_FAM_list_list[self.line_index])<(column+1)):
                        for i in range(len(self.PR_FAM_list_list[self.line_index]),column):
                            self.PR_FAM_list_list[self.line_index].append('TEM')
                        self.PR_FAM_list_list[self.line_index].append(item.text().upper())
                        QMessageBox.information(self, "通知", "该PR FAM已更新。")
                    else:
                        self.PR_FAM_list_list[self.line_index][column] = item.text().upper()
                        QMessageBox.information(self, "通知", "该PR FAM已更新。")
                if(row==1):
                    if(len(self.PR_Num_list_list[self.line_index])<(column+1)):
                        for i in range(len(self.PR_Num_list_list[self.line_index]),column):
                            self.PR_Num_list_list[self.line_index].append('TEM')
                        self.PR_Num_list_list[self.line_index].append(item.text().upper())
                        QMessageBox.information(self, "通知", "该PR Num已更新。")
                    else:
                        self.PR_Num_list_list[self.line_index][column] = item.text().upper()
                        QMessageBox.information(self, "通知", "该PR Num已更新。")
                print((self.PR_FAM_list_list[self.line_index]),(self.PR_Num_list_list[self.line_index]))

            elif (len(self.PR_Num_list_list) == 0) and (len(self.PR_FAM_list_list)==0):#都为空意味着列表为全空，需要新建PR_Num和PR_FAM
                row=item.row()
                column=item.column()
                print(row,column,item.text().upper())
                #print(len(self.PR_FAM_list_list[self.line_index]),len(self.PR_Num_list_list[self.line_index]),column)
                self.line_index=0
                PR_Num_temp = []
                PR_FAM_temp = []
                if(row==0):
                    PR_FAM_temp.append(item.text().upper())
                    self.PR_FAM_list_list.append(PR_FAM_temp)
                    QMessageBox.information(self, "通知", "该PR FAM已更新。")
                    print(len(self.PR_FAM_list_list[self.line_index]))
                if(row==1):
                    PR_Num_temp.append(item.text().upper())
                    self.PR_Num_list_list.append(PR_Num_temp)
                    QMessageBox.information(self, "通知", "该PR Num已更新。")
                    print(len(self.PR_Num_list_list[self.line_index]))
            elif (len(self.PR_Num_list_list) == 0):#只有PR_Num为空，新建PR_Num
                row=item.row()
                column=item.column()
                print(row,column,item.text().upper())
                #print(len(self.PR_FAM_list_list[self.line_index]),len(self.PR_Num_list_list[self.line_index]),column)
                self.line_index=0
                PR_Num_temp = []
                if(row==0):
                    self.PR_FAM_list_list[self.line_index].append(item.text().upper())
                    QMessageBox.information(self, "通知", "该PR FAM已更新。")
                    print(len(self.PR_FAM_list_list[self.line_index]))
                if(row==1):
                    PR_Num_temp.append(item.text().upper())
                    self.PR_Num_list_list.append(PR_Num_temp)
                    QMessageBox.information(self, "通知", "该PR Num已更新。")
                    print(len(self.PR_Num_list_list[self.line_index]))
            elif len(self.PR_FAM_list_list)==0:
                row = item.row()
                column = item.column()
                print(row, column, item.text().upper())
                self.line_index = 0
                PR_FAM_temp = []
                if (row == 1):
                    self.PR_Num_list_list[self.line_index].append(item.text().upper())
                    QMessageBox.information(self, "通知", "该PR FAM已更新。")
                    print(len(self.PR_Num_list_list[self.line_index]))
                if (row == 0):
                    PR_FAM_temp.append(item.text().upper())
                    self.PR_FAM_list_list.append(PR_FAM_temp)
                    QMessageBox.information(self, "通知", "该PR Num已更新。")
                print(len(self.PR_FAM_list_list[self.line_index]))

        #self.Display_current_PR_combination(self.line_index)

    def flash_strs(self):
        self.kopp_file_path_Edit.setText(self.kopp_file_path_str[0])

    def display_ZDC_file_Calculate_instance_output(self,message):
        self.Caculate_kopp_and_ZDC_Files_output.append(message)

    def  display_Macro_generate_instance_output(self,message):
        self.Caculate_kopp_and_ZDC_Files_output.append(message)

    def ZDC_file_Calculate_funktion(self):
        self.All_controller_data_and_programming=[]
        current_ZDC_file_caculated_data=[]
        current_ZDC_file_caculated_programming=[]
        current_ZDC_caculated_gen2_programming=[]
        print('计算所有ZDC文件')
        self.Caculate_kopp_and_ZDC_Files_output.clear()
        self.ZDC_data_output_Table.clear()
        self.ZDC_data_output_Table.setHorizontalHeaderLabels(['ID', '数据地址','数据内容',' '])
        row_count = 0
        # self.ZDC_file_Calculate_pyqtSignal.emit()
        #temp=self.ZDC_File_saved_Path+self.SG_List_result[1][2]
        #print("计算中",temp)


        if len(self.SG_List_result)==0 or len(self.PR_Num_list_list)==0:
            QMessageBox.information(self, "错误", "请先加载kopp文件，并加载ZDC status和ZDC文件所在路径")
        else:
            count=0#为了应对一个控制器对应多个ZDC文件的情况
            for index,SG_List_result_n in enumerate(self.SG_List_result):
                print('index',index,'  count', count)
                print('start')
                print('len(SG_List_result[index]',len(SG_List_result_n))
                for i in range(2,len(SG_List_result_n)):
                    print('i:',i)
                    #self.ZDC_data_output_Table.setItem(row_count, 0, SG_ID_Item)
                    if self.SG_id_and_ZDC_Table.item(index + count, 2).checkState() == Qt.Checked:
                        print('io')
                        temp = self.ZDC_File_saved_Path + self.SG_List_result[index][i]
                        print("【【【【【【【【【【【【【【【【【计算中", temp,'】】】】】】】】】】】】】】】】】】】】】】')

                        error_info,current_ZDC_file_controller_id,current_ZDC_file_caculated_data, current_ZDC_file_caculated_programming, current_ZDC_caculated_gen2_programming=self.ZDC_file_Calculate_instance.Xml_Analyse(self.PR_Num_list_list[self.line_index], temp, self.ZDC_file_path_str)

                        if error_info=='0':
                            self.SG_List_result[index][0]=current_ZDC_file_controller_id

                            SG_ID_Item = QTableWidgetItem(str(self.SG_List_result[index][0]))
                            ZDC_filename_Item = QTableWidgetItem(str(self.SG_List_result[index][i]))
                            self.ZDC_data_output_Table.setItem(row_count, 0, SG_ID_Item)
                            self.ZDC_data_output_Table.setItem(row_count, 2, ZDC_filename_Item)
                            row_count += 1

                            self.All_controller_data_and_programming.append([self.SG_List_result[index][0],self.SG_List_result[index][1],self.SG_List_result[index][i],current_ZDC_file_caculated_data,current_ZDC_file_caculated_programming,current_ZDC_caculated_gen2_programming])
                            count=count+1

                            for i in current_ZDC_file_caculated_data:
                                # print(i,PR_FAM_list.index(i),str(PR_Num_list[PR_FAM_list.index(i)]))
                                Address_Item = QTableWidgetItem(str(i[0]))
                                data_Item = QTableWidgetItem(str(i[1:]))
                                #data_Item.setFont(QFont("Times", 10, ))

                                if 'NA' in i[1:]:
                                    print('NA')
                                    data_Item.setBackground((QColor(255, 217, 0)))
                                if '**' in i[1:]:
                                    print('**')
                                    data_Item.setBackground((QColor(255, 0, 48)))

                                print('oooo')
                                self.ZDC_data_output_Table.setItem(row_count, 1, Address_Item)
                                self.ZDC_data_output_Table.setItem(row_count, 2, data_Item)
                                '''
                                if 'na' in i[1:]:
                                    print('na')
                                    self.ZDC_data_output_Table.item(row_count, 2).setBackground((QColor(0, 0, 255)))
                                if '**' in i[1:]:
                                    print('**')
                                    self.ZDC_data_output_Table.item(row_count, 2).setBackground((QColor(255, 0, 0)))
                                '''
                                row_count+=1
                                if (row_count+2)>self.ZDC_data_output_Table.rowCount():
                                    self.ZDC_data_output_Table.setRowCount(self.ZDC_data_output_Table.rowCount() + 30)#新增30行
                            if len(current_ZDC_file_caculated_programming) > 0:
                                for i in current_ZDC_file_caculated_programming:
                                    Address_Item = QTableWidgetItem("一代参数")
                                    data_Item = QTableWidgetItem(str(i))
                                    print('oooo')
                                    self.ZDC_data_output_Table.setItem(row_count, 1, Address_Item)
                                    self.ZDC_data_output_Table.setItem(row_count, 2, data_Item)

                                    row_count += 1
                                    if (row_count + 2) > self.ZDC_data_output_Table.rowCount():
                                        self.ZDC_data_output_Table.setRowCount(
                                            self.ZDC_data_output_Table.rowCount() + 30)  # 新增30行

                            if len(current_ZDC_caculated_gen2_programming) > 0:
                                for i in current_ZDC_caculated_gen2_programming:
                                    Address_Item = QTableWidgetItem("二代参数")
                                    data_Item = QTableWidgetItem(str(i))
                                    print('oooo')
                                    self.ZDC_data_output_Table.setItem(row_count, 1, Address_Item)
                                    self.ZDC_data_output_Table.setItem(row_count, 2, data_Item)

                                    row_count += 1
                                    if (row_count + 2) > self.ZDC_data_output_Table.rowCount():
                                        self.ZDC_data_output_Table.setRowCount(
                                            self.ZDC_data_output_Table.rowCount() + 30)  # 新增30行
                        else:
                            self.Caculate_kopp_and_ZDC_Files_output.append(error_info)





                    else:
                        #row_count += 1
                        count=count+1
                count=count-1


            print(self.All_controller_data_and_programming)


    def Save_ZDC_data_funktion(self):
        print("Save_ZDC_data_funktion")
        #if len(self.All_controller_data_and_programming)<1:
        #    QMessageBox.information(self, "错误", "请先计算ZDC数据，再保存ZDC数据。")
        #else:
        desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
        saved_file_path = QFileDialog.getSaveFileName(self, 'save file', desktop_path, "Excel files (*.xls);;all files(*.*)")
        if (len(saved_file_path[0])) < 6:
            return 0

        workbook = xlwt.Workbook(encoding='ascii')
        pattern = xlwt.Pattern()  # Create the Pattern
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN  # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
        pattern.pattern_fore_colour = 5  # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
        style = xlwt.XFStyle()  # Create the Pattern
        style.pattern = pattern  # Add Pattern to Style

        pattern2 = xlwt.Pattern()  # Create the Pattern
        pattern2.pattern = xlwt.Pattern.SOLID_PATTERN  # May be: NO_PATTERN, SOLID_PATTERN, or 0x00 through 0x12
        pattern2.pattern_fore_colour = 2  # May be: 8 through 63. 0 = Black, 1 = White, 2 = Red, 3 = Green, 4 = Blue, 5 = Yellow, 6 = Magenta, 7 = Cyan, 16 = Maroon, 17 = Dark Green, 18 = Dark Blue, 19 = Dark Yellow , almost brown), 20 = Dark Magenta, 21 = Teal, 22 = Light Gray, 23 = Dark Gray, the list goes on...
        style2 = xlwt.XFStyle()  # Create the Pattern
        style2.pattern = pattern2  # Add Pattern to Style

        worksheet = workbook.add_sheet('ZDC Data sheet')
        row_count=1
        worksheet.write(0, 0, QDate.currentDate() .toString(Qt.ISODate))

        for current_controller  in self.All_controller_data_and_programming :
            print("current_controller")
            controller_name = current_controller[1].replace("\n","")  # 当前控制器
            controller_id = current_controller[0]
            controller_ZDC_filename = current_controller[2]
            coding_Anpussung=current_controller[3]
            #gen1_programming=current_controller[4]
            #gen2_programming=current_controller[5]
            #print("current_controller2")
            #print(controller_id)
            #print(controller_name)
            #print(controller_ZDC_filename)
            #print(coding_Anpussung)

            worksheet.write(row_count, 0, controller_id)
            worksheet.write(row_count, 1, controller_name)
            worksheet.write(row_count, 2, controller_ZDC_filename)
            row_count+=1
            print("row")
            for data_list in current_controller[3]:  # 处理coding和Anpussung
                worksheet.write(row_count, 1, str(data_list[0]))
                if ('NA' in data_list[1:]) :
                    worksheet.write(row_count, 2, str(data_list[1:]), style)
                elif ('**' in data_list[1:]):
                    worksheet.write(row_count, 2, str(data_list[1:]), style2)
                else:
                    worksheet.write(row_count, 2, str(data_list[1:]))

                row_count += 1
            if len(current_controller[4]) > 0:  # 处理一代progranmming
                for i in current_controller[4][0:]:
                    worksheet.write(row_count, 1, '一代参数')
                    worksheet.write(row_count, 2, str(i))
                    row_count += 1

            if len(current_controller[5]) > 0:  # 处理二代progranmming
                for i in current_controller[5][0:]:
                    worksheet.write(row_count, 1, '二代参数')
                    worksheet.write(row_count, 2, str(i))
                    row_count += 1

        workbook.save(saved_file_path[0])
        QMessageBox.information(self, "通知", "数据已保存。")

    def macro_generate_funktion(self):
        self.trigger_output_main.emit(".")
        print('emited')
        '''
        if len(self.All_controller_data_and_programming)<1:
            QMessageBox.information(self, "错误", "请先计算ZDC数据，然后再生成Macro。")
        else:
            #self.Caculate_kopp_and_ZDC_Files_output.clear()
            self.Macro_generate_instance.Macro_generate_function(self.SG_List_result,self.All_controller_data_and_programming)
            QMessageBox.information(self, "通知", "Macro已经生成在桌面上。")
        '''



class Selection_Dialog(QDialog):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.select_all_checkbox_index=0
        self.all_programming_and_data_for_selection=[]
        self.Macro_generate_instance_in_selection=Macro_generate_class()#初始化


        self.resize(500, 700)
        print('1')
        # 表格布局，用来布局QLabel和QLineEdit及QSpinBox
        grid = QGridLayout()
        grid.setSpacing(10)
        print('2')

        self.SG_Selection_Table = QTableWidget(30,3)
        self.SG_Selection_Table.setHorizontalHeaderLabels(['ID','描述','ZDC文件名'])
        self.SG_Selection_Table.verticalHeader().setHidden(True);
        grid.addWidget(self.SG_Selection_Table,1,1,8,6)
        self.SG_Selection_Table.setColumnWidth(0, 50)
        self.SG_Selection_Table.setColumnWidth(1, 150)
        self.SG_Selection_Table.setColumnWidth(2, 300)

        self.select_all_checkbox = QCheckBox('全选', self)
        self.select_all_checkbox.stateChanged.connect(self.select_all_checkbox_changed)
        grid.addWidget(self.select_all_checkbox, 10, 1,1,1)


        # 创建ButtonBox，用户确定和取消
        buttonBox = QDialogButtonBox(parent=self)
        buttonBox.setOrientation(Qt.Horizontal) # 设置为水平方向
        buttonBox.setStandardButtons(QDialogButtonBox.Cancel|QDialogButtonBox.Ok) # 确定和取消两个按钮
        # 连接信号和槽
        buttonBox.accepted.connect(self.Macro_generation_confirmed) # 确定
        buttonBox.rejected.connect(self.close) # 取消
        print('4')
        # 垂直布局，布局表格及按钮
        layout = QVBoxLayout()
        # 加入前面创建的表格布局
        layout.addLayout(grid)

        # 放一个间隔对象美化布局
        #spacerItem = QSpacerItem(20, 48, QSizePolicy.Minimum, QSizePolicy.Expanding)
        #layout.addItem(spacerItem)
        print('5')
        # ButtonBox
        layout.addWidget(buttonBox)

        self.setLayout(layout)
        self.show()


    def display_macro_selection_Funktion(self,all_programming_and_data):
        print("dialog,display_macro_selection_Funktion")
        self.all_programming_and_data_for_selection=all_programming_and_data
        #self.leName.setText(all_programming_and_data[0][2])
        if len(all_programming_and_data)<1:
            return
        else:
            if len(all_programming_and_data) > self.SG_Selection_Table.rowCount():
                self.SG_Selection_Table.setRowCount( len(all_programming_and_data) + 10)  # 需要统计SG_List_result列表中第三个元素的总和作为总行数，后期再优化
            print('调整长度')
            for index, all_programming_and_data_n in enumerate(all_programming_and_data):
                SG_id_Item = QTableWidgetItem(str(all_programming_and_data_n[0]))
                SG_besch_Item = QTableWidgetItem(str(all_programming_and_data_n[1]))
                self.SG_Selection_Table.setItem(index, 0, SG_id_Item)
                self.SG_Selection_Table.setItem(index, 1, SG_besch_Item)

                SG_ZDC_filename_Item = QTableWidgetItem(str(all_programming_and_data_n[2]))
                checkBox = QTableWidgetItem(SG_ZDC_filename_Item)
                checkBox.setCheckState(Qt.Checked)
                self.SG_Selection_Table.setItem(index, 2, checkBox)


    def select_all_checkbox_changed(self):
        if (self.select_all_checkbox_index==0):
            self.select_all_checkbox_index=1
            for index, all_programming_and_data_n in enumerate(self.all_programming_and_data_for_selection):
                SG_ZDC_filename_Item = QTableWidgetItem(str(all_programming_and_data_n[2]))
                checkBox = QTableWidgetItem(SG_ZDC_filename_Item)
                checkBox.setCheckState(Qt.Unchecked)
                self.SG_Selection_Table.setItem(index, 2, checkBox)
        else:
            self.select_all_checkbox_index = 0
            for index, all_programming_and_data_n in enumerate(self.all_programming_and_data_for_selection):
                SG_ZDC_filename_Item = QTableWidgetItem(str(all_programming_and_data_n[2]))
                checkBox = QTableWidgetItem(SG_ZDC_filename_Item)
                checkBox.setCheckState(Qt.Checked)
                self.SG_Selection_Table.setItem(index, 2, checkBox)

    def Macro_generation_confirmed(self):
        Macro_tobe_generated_data_and_programming=[]
        print('Macro_generation_confirmed')
        for index, all_programming_and_data_n in enumerate(self.all_programming_and_data_for_selection):
            if self.SG_Selection_Table.item(index, 2).checkState() == Qt.Checked:
                Macro_tobe_generated_data_and_programming.append(self.all_programming_and_data_for_selection[index])
        print(len(Macro_tobe_generated_data_and_programming))
        print(Macro_tobe_generated_data_and_programming)
        if len(Macro_tobe_generated_data_and_programming)>0:
            flag_saved=self.Macro_generate_instance_in_selection.Macro_generate_function([],Macro_tobe_generated_data_and_programming)
            if flag_saved:
                QMessageBox.information(self, "通知", "Macro已生成。\n请将Macro文件拷贝至C:\Program Files (x86)\Offboard_Diagnostic_Information_System_Engineering\macros文件夹下使用。\n如果使用其他电脑写入，请将桌面ProjektZ文件夹下的Programming_Files文件夹一起拷贝过去。")

class registration_Dialog(QDialog):
    trigger_output_r = PyQt5.QtCore.pyqtSignal(str)


    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        self.resize(500, 200)
        print('1')
        # 表格布局，用来布局QLabel和QLineEdit及QSpinBox
        self.setWindowTitle('ZDC 自动计算与注入系统 （请注册后使用）')
        self.setWindowIcon(QIcon("icons/logo.ico"))
        grid = QGridLayout()
        grid.setSpacing(10)
        print('2')
        self.Label_C_panel = QLabel('请将该数字序列发给管理员(EPEI-1)获取激活码')
        self.Line_C_panel = QLineEdit(self)
        self.Label_encrypted_C_panel = QLabel('请输入激活码')
        self.Line_encrypted_C_panel = QLineEdit(self)
        grid.addWidget(self.Label_C_panel, 1, 1,1,3)
        grid.addWidget(self.Line_C_panel, 1, 5,1,3)
        grid.addWidget(self.Label_encrypted_C_panel, 2, 1,1,3)
        grid.addWidget(self.Line_encrypted_C_panel, 2, 3,1,5)

        CVolumeSerialNumber = win32api.GetVolumeInformation("C:\\")[1]
        self.Line_C_panel.setText(str(CVolumeSerialNumber))
        print("请将该字符串发给管理员获取激活码：", CVolumeSerialNumber)
        k = des("DESCRYPT", CBC, "\1\0\1\0\0\1\0\0", pad=None, padmode=PAD_PKCS5)
        self.d_calcu = k.encrypt(str(CVolumeSerialNumber))


        # 创建ButtonBox，用户确定和取消
        buttonBox = QDialogButtonBox(parent=self)
        buttonBox.setOrientation(Qt.Horizontal) # 设置为水平方向
        buttonBox.setStandardButtons(QDialogButtonBox.Cancel|QDialogButtonBox.Ok) # 确定和取消两个按钮
        # 连接信号和槽
        buttonBox.accepted.connect(self.Compare_registration_code) # 确定
        buttonBox.rejected.connect(self.close) # 取消
        print('4')
        # 垂直布局，布局表格及按钮
        layout = QVBoxLayout()
        # 加入前面创建的表格布局
        layout.addLayout(grid)

        # 放一个间隔对象美化布局
        #spacerItem = QSpacerItem(20, 48, QSizePolicy.Minimum, QSizePolicy.Expanding)
        #layout.addItem(spacerItem)
        print('5')
        # ButtonBox
        layout.addWidget(buttonBox)
        self.setLayout(layout)
        self.show()



    def Compare_registration_code(self):
        print("Compare_registration_code")
        d_input = self.Line_encrypted_C_panel.text()

        if str(self.d_calcu) == str(d_input):
            print("Ture")

            desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
            print(desktop_path)
            License_file_path = desktop_path + "\\ProjektZ\\"
            if (os.path.exists(License_file_path)):
                pass
            else:
                os.makedirs(License_file_path)
            print('55')
            License_filename = License_file_path + "License.z"
            f = open(License_filename, 'w')  # 打开文件
            f.write(str(d_input))  # 写数据地址到文件中
            f.close()


            self.trigger_output_r.emit("io")
            self.close()

        else:
            self.Line_encrypted_C_panel.setText("输入错误，请检查后再尝试。")


class Main_Window(QMainWindow):

    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):

        self.statusBar()
        openFile = QAction(QIcon('open.png'), '帮助', self)
        #openFile.setShortcut('Ctrl+O')
        #openFile.setStatusTip('Open new File')
        openFile.triggered.connect(self.showDialog)
        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&文件')
        fileMenu.addAction(openFile)

        self.ZDC_file_locate_Widget_instance=ZDC_file_locate_Widget()
        self.ZDC_file_locate_Widget_instance.trigger_output_main.connect(self.display_macro_selection_dialog)

        #self.PR_Num_Calculate_Widget=PR_Num_Calculate_Widget()
        #self.ZDC_file_Calculate_and_macro_generate_Widget=ZDC_file_Calculate_and_macro_generate_Widget()

        #self.ZDC_file_Calculate_and_macro_generate_Widget.ZDC_file_Calculate_pyqtSignal.connect(self.PR_Num_Calculate_Widget.return_PR_Num)
        #self.PR_Num_Calculate_Widget.return_PR_Num_pyqtSignal.connect(self.ZDC_file_Calculate_and_macro_generate_Widget.recieve_PR_Num)

        main_window_widget = QWidget(self)
        self.grid = QGridLayout(main_window_widget)
        self.grid.setSpacing(10)
        self.grid.addWidget(self.ZDC_file_locate_Widget_instance, 1, 1)
        #self.grid.addWidget(self.PR_Num_Calculate_Widget, 1, 2)
        #self.grid.addWidget(self.ZDC_file_Calculate_and_macro_generate_Widget, 2,1,2,2)
        self.setCentralWidget(main_window_widget)



        self.setGeometry(100, 100, 1550, 850)
        self.setWindowTitle('ZDC 自动计算与注入系统 （公测版） V1.2')
        self.setWindowIcon(QIcon("icons/logo.ico"))
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")
        '''
        tuopan = QSystemTrayIcon(main_window_widget)
        icon1 = QIcon("icons/logo.jpg")
        tuopan.setIcon(icon1)
        tuopan.show()
        '''
        desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
        License_file_path = desktop_path + "\\ProjektZ\\"
        License_filename = License_file_path + "License.z"

        CVolumeSerialNumber_main = win32api.GetVolumeInformation("C:\\")[1]
        k_main = des("DESCRYPT", CBC, "\1\0\1\0\0\1\0\0", pad=None, padmode=PAD_PKCS5)
        self.d_calcu_main = k_main.encrypt(str(CVolumeSerialNumber_main))

        if (os.path.isfile(License_filename)):
            f = open(License_filename, 'r')  # 打开文件
            licensefile_text = f.read()  # 写数据地址到文件中
            f.close()
            # print(licensefile_text)
            if (str(licensefile_text) == str(self.d_calcu_main)):
                print('Ture')
                #self.show()
                desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
                Real_log_file_path = desktop_path.replace('Desktop', '') + "PyProject\\"
                Real_log_filename = Real_log_file_path + "Log.txt"
                print('Real_log_file_path',Real_log_file_path)
                if (os.path.exists(Real_log_file_path)):
                    pass
                else:
                    os.makedirs(Real_log_file_path)

                now = QDate.currentDate()
                if (now.toString(Qt.ISODate)).split('-')[0]!='2018' and (now.toString(Qt.ISODate)).split('-')[0]!='2017':
                    print(os.path.exists(Real_log_filename))
                    QMessageBox.information(self, "错误", "log文件错误。")
                    if (os.path.exists(Real_log_filename)):
                        pass
                    else:
                        f = open(Real_log_filename, 'w')
                        f.write('1')
                        f.close()
                else:
                    if (os.path.exists(Real_log_filename)):
                        #f = open(Real_log_filename, 'x')
                        #if str(f.read())=='1':
                        #    f.close()
                        QMessageBox.information(self, "错误","log文件错误。")
                    else:
                        self.showMaximized()
                        print('MAXmized')
            else:
                self.registration_Dialog_instance = registration_Dialog()
                #self.registration_Dialog_instance.trigger_output_r.connect(self.show)
                self.registration_Dialog_instance.trigger_output_r.connect(self.showMaximized)
        else:
            self.registration_Dialog_instance = registration_Dialog()
            #self.registration_Dialog_instance.trigger_output_r.connect(self.show)
            self.registration_Dialog_instance.trigger_output_r.connect(self.showMaximized)

    def showDialog(self):
        '''fname = QFileDialog.getOpenFileName(self, 'Open file', '/home')
        if fname[0]:
            f = open(fname[0], 'r')
            with f:
                data = f.read()
                self.ZDC_file_locate_Widget.ZDC_status_Edit.setText(data)
                '''
        QMessageBox.information(self, "帮助", "      1.加载ZDC Status文件、ZDC文件路径，诊断报告可以加载也可以不加载， 点击自动生成ZDC列表按钮。\n\
       2.加载Kopp文件，点击解析kopp文件按钮，并选择装备条线。\n\
       3.勾选要计算的ZDC文件，然后点击计算ZDC数据按钮。\n\
       4.点击生成Macro文件按钮，勾选要写入的控制器，然后点击OK生成Macro。\n\
       5.有合作或问题请联系zhang126biao@126.com")

    def display_macro_selection_dialog(self):
        print("main, display_macro_selection_dialog")
        if len(self.ZDC_file_locate_Widget_instance.All_controller_data_and_programming)<1:
            QMessageBox.information(self, "错误", "请先计算ZDC数据，再生成Macro。")
        else:
            self.dialog_instance = Selection_Dialog()
            self.dialog_instance.display_macro_selection_Funktion(self.ZDC_file_locate_Widget_instance.All_controller_data_and_programming)
            #self.dialog_instance.destroy()


#if __name__ == '__main__':
app = QApplication(sys.argv)

#ex = ZDC_file_locate_Widget()
ex = Main_Window()
#ex2=PR_Num_Calculate_Widget()
sys.exit(app.exec_())




