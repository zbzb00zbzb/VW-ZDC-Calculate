
import operator
from array import array
from openpyxl import Workbook
from openpyxl import load_workbook
import operator
from array import array
import re
import shutil
import gzip, io
import xml.dom.minidom
from xml.dom.minidom import parse
from functools import reduce
#from Excel import *
from bs4 import BeautifulSoup
import bs4
import os,os.path
import shutil
import sys
import PyQt5.QtCore
from PyQt5.QtCore import pyqtSignal, QObject
from PyQt5.QtWidgets import (QMainWindow, QTextEdit,
                             QAction, QFileDialog, QApplication, QLabel, QLineEdit, QGridLayout, QWidget, QPushButton,QMessageBox)
from PyQt5.QtGui import QIcon





class Macro_generate_class(QWidget):
    trigger_output_M = PyQt5.QtCore.pyqtSignal(str)

    def __init__(self):
        print('Macro_generate_class chushihua ')

        super().__init__()

        self.initUI()

    def initUI(self):
        self.PR_Num_list=[]
        self.ZDC_file_list=[]

    def example(self,text):
        print('example')
        self.trigger_output_Z.emit(text)


    def Macro_generate_function(self,SG_List_result,All_controller_data_and_programming):
        trigger_output_M = PyQt5.QtCore.pyqtSignal(str)
        desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
        #fname = QFileDialog.getOpenFileName(self, 'Open file','.',"xlsx files (*.xlsx *.xls)")
        saved_file_path =  QFileDialog.getSaveFileName(self,'save file',desktop_path ,"Python files (*.py);;all files(*.*)")
        if(len(saved_file_path[0]))<6:
            return 0
        print(saved_file_path[0])
        if(os.path.exists(saved_file_path[0])):
            pass
        else:
            print('创建文件')
            #os.mknod(saved_file_path[0])
        filename=saved_file_path[0]
        f = open(filename, 'w')
        print('创建完成')
        #f = open(saved_file_path[0], 'w')

        '''
                for i, current_SG in enumerate(SG_List_result):

                    controller_name = current_SG[1] # 当前控制器
                    controller_id = current_SG[0]
                    f.write('\nprint "Wechsel der Diagnosesitzung wird durchgefuhrt: ')
                    f.write(controller_name)
                    f.write('"\nresultConnectToEcu = diagnosticInterface.connectToEcu(0x')
                    f.write(controller_id)
                    f.writelines([')\n', 'diagnosticInterface.openConnection(resultConnectToEcu.getConnectionHandle())\n', \
                                  'diagnosticInterface.switchSession(resultConnectToEcu.getConnectionHandle(), '])
                    if operator.eq(controller_id, '09'):
                        f.write('"DiagnServi_DiagnSessiContrDevelSessi")\n')
                    else:
                        f.write('"DiagnServi_DiagnSessiContrVWEndOfLineSessi")\n')
                    f.write('print "Zugriffberechtigung wird durchgefuhrt: ')
                    f.write(controller_name)
                    f.writelines(['"\n', 'diagnosticInterface.securityAccess(resultConnectToEcu.getConnectionHandle(), "'])
                    if operator.eq(controller_id, '09'):
                        f.write('31347')
                    else:
                        f.write('20103')
                    f.writelines(
                        ['", "Login")\n', 'print "Hex-Services werden geschrieben: ', controller_name, '"\n'])
        '''

        f.writelines(['# coding: latin-1\n',\
        'import sys\n',\
        'import operator\n',\
        'from array import array\n',\
        'import base64\n',\
        '#sys.setdefaultencoding( "latin-1")\n',\
        'from java.lang import Boolean\n',\
        'from java.math import BigInteger\n',\
        'from java.util import HashMap\n',\
        'from java.util import ArrayList\n',\
        'from de.volkswagen.odis.vaudas.vehiclefunction.automation import IDiagnosticInterface\n',\
        'from de.volkswagen.odis.vaudas.vehiclefunction.automation.types import IDiagResultConnectEcu\n',\
        '#from de.volkswagen.odis.vaudas.vehiclefunction.automation import IEcuStateInterface\n',\
        'diagnosticInterface = IDiagnosticInterface.Factory.getInstance()\n',\
        'diagnosticInterface.configureSetting("Multilink.MaxNumberOfLogicalLinks", "1")\n',\
        'diagnosticInterface.startProtocol()\n'])
        '''
        'sys.setdefaultencoding( "latin-1")\n',\ ODIS 7,2,2要删掉这句
        然后将        
        'from de.volkswagen.odis.vaudas.vehiclefunction.automation import IEcuStateInterface\n',\
        换成
        'from de.volkswagen.odis.vaudas.vehiclefunction.automation import ITotalSystemsInterface\n',\     
        '''
        for current_controller  in All_controller_data_and_programming :
            controller_name = current_controller[1].replace("\n","")
            controller_id = current_controller[0]
            f.write('\nprint "Wechsel der Diagnosesitzung wird durchgefuhrt: ')
            f.write(controller_name)
            f.write('"\nresultConnectToEcu = diagnosticInterface.connectToEcu(0x')
            f.write(controller_id)
            f.writelines([')\n', 'diagnosticInterface.openConnection(resultConnectToEcu.getConnectionHandle())\n', \
                          'diagnosticInterface.switchSession(resultConnectToEcu.getConnectionHandle(), '])
            if operator.eq(controller_id, '09'):
                f.write('"DiagnServi_DiagnSessiContrDevelSessi")\n')
            else:
                f.write('"DiagnServi_DiagnSessiContrVWEndOfLineSessi")\n')

            if len(current_controller[4]) > 0:
                for i in current_controller[4][0:]:
                    f.write('print(\"Writing gen-1 Programming file:')
                    f.write(str(i).split('\\')[-1])
                    f.write('")\n')

                    f.writelines(['prr1= diagnosticInterface.sendRawService(resultConnectToEcu.getConnectionHandle(), "22 04 0A")\n'])
                    ######写入C盘路径
                    f.write('diagnosticInterface.dataSetDownload(resultConnectToEcu.getConnectionHandle(), ')
                    # f.write(str(i))
                    #f.write("sys.path[0]+\"\\\\Programming_Files\\\\")
                    f.write("\"C:\\\\Program Files (x86)\\\\Offboard_Diagnostic_Information_System_Engineering\\\\macros\\\\Programming_Files\\\\")
                    f.write(str(i).split('\\')[-1])
                    f.write('")\n')
                    ###写入桌面路径
                    f.write('diagnosticInterface.dataSetDownload(resultConnectToEcu.getConnectionHandle(), ')
                    temp = str(i).split('\\')
                    f.write("\"C:")
                    for i in temp[1:]:
                        f.write("\\\\")
                        f.write(i)
                    f.write('")\n')
                    ###判断是否成功写入
                    f.writelines(['prr2= diagnosticInterface.sendRawService(resultConnectToEcu.getConnectionHandle(), "22 04 0A")\n'])
                    f.write('if prr2[4]-prr1[4] >= 1:\n    print(\'Programming file successfully writen.\')\nelse:\n    print(\'!!! !!! !!!Programming file writen failed, please try writing mannually!!! !!! !!! \')\n')

            if len(current_controller[5]) > 0:
                for i in current_controller[5][0:]:
                    f.write('print(\"Writing gen-2 Programming file:')
                    f.write(str(i).split('\\')[-1])
                    f.write('")\n')

                    f.write('diagnosticInterface.dataSetDownload(resultConnectToEcu.getConnectionHandle(), ')
                    # f.write(str(i))
                    # f.write("sys.path[0]+\"\\\\Programming_Files\\\\")
                    f.write("\"C:\\\\Program Files (x86)\\\\Offboard_Diagnostic_Information_System_Engineering\\\\macros\\\\Programming_Files\\\\")
                    f.write(str(i).split('\\')[-1])
                    '''
                    temp = str(i).split('\\')
                    f.write("\"C:")
                    for i in temp[1:]:
                        f.write("\\\\")
                        f.write(i)
                    '''

                    f.write('")\n')
            #f.write('print "Zugriffberechtigung wird durchgefuhrt: ')
            #f.write(controller_name)
            #f.write('"\n')

            if operator.eq(controller_id, '09'):
                f.write("code=\"MzEzNDc=\"\n")#31347
                f.write("controller_pd= base64.b64decode(code)\n")
                f.write('diagnosticInterface.securityAccess(resultConnectToEcu.getConnectionHandle(), ')
                f.write('controller_pd')
                f.writelines([', "Login")\n'])
            elif(operator.eq(controller_id, '03')):
                f.write("code=\"MjA3OTU=\"\n")#20795
                f.write("controller_pd= base64.b64decode(code)\n")
                f.write('diagnosticInterface.securityAccess(resultConnectToEcu.getConnectionHandle(), ')
                f.write('controller_pd')
                f.writelines([', "Login")\n'])
            elif (operator.eq(controller_id, '44')):
                f.write("code=\"MTkyNDk=\"\n")#19249
                f.write("controller_pd= base64.b64decode(code)\n")
                f.write('diagnosticInterface.securityAccess(resultConnectToEcu.getConnectionHandle(), ')
                f.write('controller_pd')
                f.writelines([', "Login")\n'])
            elif (operator.eq(controller_id, '01'))or (operator.eq(controller_id, '08'))or (operator.eq(controller_id, 'A5')) or (operator.eq(controller_id, '16')) or (operator.eq(controller_id, '17'))or (operator.eq(controller_id, '02')):
                pass
            else:
                f.write("code=\"MjAxMDM=\"\n")#20103
                f.write("controller_pd= base64.b64decode(code)\n")
                f.write('diagnosticInterface.securityAccess(resultConnectToEcu.getConnectionHandle(), ')
                f.write('controller_pd')
                f.writelines([', "Login")\n'])

            f.writelines(['print "Hex-Services werden geschrieben: ', controller_name, '"\n'])

            for data_list in current_controller[3]:

                if(len(data_list[0])==4):#判断地址是不是4位
                    #print(data_list[0][0:1],controller_id)
                    #print(operator.eq(data_list[0][0:2], '72'),operator.eq(controller_id, '6D'))
                    if ('NA' in data_list[1:]) or ('**' in data_list[1:]):
                        print('该地址有NA数据，将会被跳过：',data_list[0])
                        continue
                    elif(operator.eq(data_list[0], '7200')):
                        continue
                    elif(operator.eq(data_list[0][0:2], '72')) and operator.eq(controller_id, '6D'):
                        continue
                    else:
                        #for real_data in data_list[1:]:
                        f.writelines([ 'arr1= diagnosticInterface.sendRawService(resultConnectToEcu.getConnectionHandle(), "2E '])
                        middle=data_list[0]
                        print(middle)
                        print(middle[0:2])
                        f.write(middle[0:2])
                        f.write(' ')
                        f.write(middle[2:4])

                        #middle2=list(map(int(), ws.cell(row, 5).value.split()))
                        #middle2=list(int(n for n in ws.cell(row, 5).value.split(),base=10))
                        print('1111')
                        middle3=[int(item,16) for item in data_list[1:]]
                        #print(middle3)
                        arr1=array('f',middle3)
                        #print(arr1.tolist()==middle3)
                        print('middle3:',middle3)
                        for i in data_list[1:]:
                            f.write(' ')
                            f.write(str(i))
                        f.write('")\n')
                        f.writelines([ 'arr2= diagnosticInterface.sendRawService(resultConnectToEcu.getConnectionHandle(), "22 '])
                        f.write(middle[0:2])
                        f.write(' ')
                        f.write(middle[2:4])
                        f.write('")\n')
                        f.write('arr3 = array(\'b\', [98, ')
                        print('2222')
                        if int(middle[0:2],16)>127:
                            f.write(str(int(middle[0:2], 16)-256))
                        else:
                            f.write(str(int(middle[0:2], 16)))
                        f.write(', ')
                        if int(middle[2:4],16)>127:
                            f.write(str(int(middle[2:4], 16)-256))
                        else:
                            f.write(str(int(middle[2:4], 16)))
                        f.write(', ')
                        for i in middle3:
                            if i>127:
                                i=i-256
                            f.write(str(i))
                            f.write(', ')
                        f.write('])\n')
                        f.writelines(['if arr1[0] == 110:\n','    print(\'Address '])
                        f.write(middle[0:2])
                        f.write(middle[2:4])
                        f.write(' ')
                        f.write( ' successfully writen.\')\nelse: \n    print(\'!!! !!! !!! Address ')
                        f.write(middle[0:2])
                        f.write(middle[2:4])
                        f.write(' ')
                        f.write( 'writen failed, please try writing mannually!!! !!! !!! \')\n')

                        f.write('if arr2 == arr3:\n    print(\'Address ')
                        f.write(middle[0:2])
                        f.write(middle[2:4])
                        f.write(' ')
                        f.write( ' successfully compared.\')\nelse: \n    print(\'!!! !!! !!!  Address ')
                        f.write(middle[0:2])
                        f.write(middle[2:4])
                        f.write(' ')
                        f.write( 'comparing failed, please check!!! !!! !!! \')\n\n')
                        print('3333')
                else:
                    #print(str(ws.cell(row,2).value))
                    pass

        f.close()
        return 1